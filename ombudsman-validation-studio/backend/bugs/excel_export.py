"""
Excel Export Service for Bug Reports

Generates professional Excel workbooks with bug report data,
including summary statistics, detailed bug listings, and sample data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional
import json
from datetime import datetime

from .models import BugReport, Bug


class BugReportExcelExporter:
    """Export bug reports to Excel format with professional formatting"""

    def __init__(self):
        # Define color scheme
        self.colors = {
            'header_bg': 'B4C7E7',        # Blue header background
            'critical': 'FF0000',         # Red
            'high': 'FFA500',             # Orange
            'medium': 'FFFF00',           # Yellow
            'low': 'ADD8E6',              # Light Blue
            'info': 'D3D3D3',             # Light Gray
            'approved': '90EE90',         # Light Green
            'rejected': 'FFB6C1',         # Light Pink
            'created': '32CD32',          # Lime Green
            'failed': 'DC143C'            # Crimson
        }

    def generate_excel_report(
        self,
        report: BugReport,
        output_path: Optional[Path] = None
    ) -> Path:
        """
        Generate a comprehensive Excel report.

        Args:
            report: BugReport object to export
            output_path: Optional output file path

        Returns:
            Path to the generated Excel file
        """
        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Add sheets
        self._create_summary_sheet(wb, report)
        self._create_bugs_sheet(wb, report)

        # Add sample data sheets for bugs with samples
        bugs_with_samples = [bug for bug in report.bugs if bug.sample_data]
        for idx, bug in enumerate(bugs_with_samples[:5], 1):  # Limit to 5 sample sheets
            self._create_sample_data_sheet(wb, bug, idx)

        # Determine output path
        if output_path is None:
            export_dir = Path("bug_reports") / "exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            output_path = export_dir / f"{report.report_id}.xlsx"

        # Save workbook
        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, wb: Workbook, report: BugReport):
        """Create summary sheet with report overview"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = 'Bug Report Summary'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Report Information
        row = 3
        info_fields = [
            ('Report ID:', report.report_id),
            ('Project:', report.project_name),
            ('Batch Job:', report.batch_job_name),
            ('Generated:', report.summary.generated_at.strftime('%Y-%m-%d %H:%M:%S')),
            ('Total Bugs:', report.summary.total_bugs),
            ('Submitted to Azure:', 'Yes' if report.submitted_to_azure else 'No')
        ]

        for label, value in info_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Severity Breakdown
        row += 1
        ws[f'A{row}'] = 'Severity Breakdown'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

        severity_data = [
            ('Critical', report.summary.by_severity.get('critical', 0), self.colors['critical']),
            ('High', report.summary.by_severity.get('high', 0), self.colors['high']),
            ('Medium', report.summary.by_severity.get('medium', 0), self.colors['medium']),
            ('Low', report.summary.by_severity.get('low', 0), self.colors['low']),
            ('Info', report.summary.by_severity.get('info', 0), self.colors['info'])
        ]

        for severity, count, color in severity_data:
            ws[f'A{row}'] = severity
            ws[f'B{row}'] = count
            ws[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            row += 1

        # Status Breakdown
        row += 1
        ws[f'A{row}'] = 'Status Breakdown'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

        status_data = [
            ('Pending Review', report.summary.by_status.get('pending_review', 0)),
            ('Approved', report.summary.by_status.get('approved', 0)),
            ('Rejected', report.summary.by_status.get('rejected', 0)),
            ('Created in Azure', report.summary.by_status.get('created_in_azure', 0)),
            ('Failed to Create', report.summary.by_status.get('failed_to_create', 0))
        ]

        for status, count in status_data:
            ws[f'A{row}'] = status
            ws[f'B{row}'] = count
            row += 1

        # Category Breakdown
        row += 1
        ws[f'A{row}'] = 'Category Breakdown'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

        for category, count in report.summary.by_category.items():
            ws[f'A{row}'] = category.replace('_', ' ').title()
            ws[f'B{row}'] = count
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

    def _create_bugs_sheet(self, wb: Workbook, report: BugReport):
        """Create detailed bugs listing sheet"""
        ws = wb.create_sheet("Bugs")

        # Headers
        headers = [
            'Bug ID', 'Title', 'Severity', 'Category', 'Status',
            'Step Name', 'Validation Type', 'Table', 'Column',
            'Expected Value', 'Actual Value', 'Failure Count',
            'Error Message', 'Work Item ID', 'Created At'
        ]

        # Write headers with styling
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header_bg'],
                                   end_color=self.colors['header_bg'],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(style='medium', color='000000')
            )

        # Write bug data
        for row_idx, bug in enumerate(report.bugs, 2):
            row_data = [
                bug.bug_id,
                bug.title,
                bug.severity.upper(),
                bug.category.replace('_', ' ').title(),
                bug.status.replace('_', ' ').title(),
                bug.step_name,
                bug.validation_type,
                bug.table_name or '',
                bug.column_name or '',
                bug.expected_value or '',
                bug.actual_value or '',
                bug.failure_count or '',
                bug.error_message or '',
                bug.work_item_id or '',
                bug.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Color-code severity column
                if col_idx == 3:  # Severity column
                    severity = bug.severity.lower()
                    if severity in self.colors:
                        cell.fill = PatternFill(
                            start_color=self.colors[severity],
                            end_color=self.colors[severity],
                            fill_type='solid'
                        )

                # Color-code status column
                if col_idx == 5:  # Status column
                    status_key = bug.status.lower().replace('_', '')
                    if 'approved' in status_key:
                        color = self.colors['approved']
                    elif 'rejected' in status_key:
                        color = self.colors['rejected']
                    elif 'created' in status_key:
                        color = self.colors['created']
                    elif 'failed' in status_key:
                        color = self.colors['failed']
                    else:
                        color = None

                    if color:
                        cell.fill = PatternFill(
                            start_color=color,
                            end_color=color,
                            fill_type='solid'
                        )

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(header)

            # Find max content length in column
            for row_idx in range(2, len(report.bugs) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set width (with max cap)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Enable auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # Freeze top row
        ws.freeze_panes = "A2"

    def _create_sample_data_sheet(self, wb: Workbook, bug: Bug, sheet_number: int):
        """Create sample data sheet for a bug"""
        if not bug.sample_data or len(bug.sample_data) == 0:
            return

        # Sheet name (truncate bug ID if too long)
        bug_id_short = bug.bug_id[:20] if len(bug.bug_id) > 20 else bug.bug_id
        sheet_name = f"Sample_{sheet_number}_{bug_id_short}"
        ws = wb.create_sheet(sheet_name)

        # Title
        ws['A1'] = f'Sample Data for Bug: {bug.bug_id}'
        ws['A1'].font = Font(size=12, bold=True)
        ws.merge_cells('A1:E1')

        # Bug context
        ws['A2'] = f'Table: {bug.table_name or "N/A"}'
        ws['A3'] = f'Step: {bug.step_name}'
        ws['A4'] = f'Validation: {bug.validation_type}'

        # Sample data headers
        row = 6
        sample = bug.sample_data[0]
        headers = list(sample.keys())

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header_bg'],
                                   end_color=self.colors['header_bg'],
                                   fill_type='solid')

        # Sample data rows (limit to 100 rows)
        for sample_idx, sample_row in enumerate(bug.sample_data[:100], row + 1):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=sample_idx, column=col_idx)
                value = sample_row.get(header, '')
                cell.value = value

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(header))

            for sample_idx in range(row + 1, min(row + 101, len(bug.sample_data) + row + 1)):
                cell_value = ws.cell(row=sample_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # Enable auto-filter
        ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"

        # Freeze header row
        ws.freeze_panes = f"A{row + 1}"


# Singleton instance
excel_exporter = BugReportExcelExporter()
