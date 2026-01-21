from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
import os
import json
from typing import List, Dict, Any
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path
import io

# Export libraries
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

RESULTS_DIR = "results"
RETENTION_DAYS = 7  # Keep results for 7 days
BASELINE_FILE = "results/.baseline.json"  # File to store current baseline

# Ensure results directory exists
os.makedirs(RESULTS_DIR, exist_ok=True)


def cleanup_old_results():
    """
    Cleanup result files older than RETENTION_DAYS.
    This runs automatically when results are fetched.
    """
    try:
        if not os.path.exists(RESULTS_DIR):
            return

        cutoff_date = datetime.now() - timedelta(days=RETENTION_DAYS)
        deleted_count = 0

        for file in os.listdir(RESULTS_DIR):
            if not file.endswith(".json"):
                continue

            file_path = os.path.join(RESULTS_DIR, file)

            try:
                # Check file modification time
                file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))

                if file_mtime < cutoff_date:
                    # Also check the timestamp inside the file to be sure
                    try:
                        with open(file_path, 'r') as f:
                            data = json.load(f)
                            timestamp_str = data.get('started_at') or data.get('timestamp') or data.get('execution_time')

                            if timestamp_str:
                                try:
                                    file_timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                                    if file_timestamp.replace(tzinfo=None) < cutoff_date:
                                        Path(file_path).unlink()
                                        deleted_count += 1
                                        print(f"[CLEANUP] Deleted old result file: {file} (age: {(datetime.now() - file_timestamp.replace(tzinfo=None)).days} days)")
                                except:
                                    # If timestamp parsing fails, use file mtime
                                    Path(file_path).unlink()
                                    deleted_count += 1
                                    print(f"[CLEANUP] Deleted old result file: {file} (based on file mtime)")
                            else:
                                # No timestamp in file, use file mtime
                                Path(file_path).unlink()
                                deleted_count += 1
                                print(f"[CLEANUP] Deleted old result file: {file} (no timestamp found)")
                    except json.JSONDecodeError:
                        # If file is corrupted, delete it anyway
                        Path(file_path).unlink()
                        deleted_count += 1
                        print(f"[CLEANUP] Deleted corrupted result file: {file}")

            except Exception as e:
                print(f"[CLEANUP] Error processing {file}: {e}")

        if deleted_count > 0:
            print(f"[CLEANUP] Cleaned up {deleted_count} old result file(s)")

    except Exception as e:
        print(f"[CLEANUP] Cleanup error: {e}")


@router.get("")
def fetch_results():
    """
    Fetch all validation results from the results directory.
    Automatically cleans up results older than RETENTION_DAYS.
    """
    # Run cleanup before fetching results
    cleanup_old_results()

    entries = []
    if os.path.exists(RESULTS_DIR):
        for file in os.listdir(RESULTS_DIR):
            if file.endswith(".json"):
                try:
                    with open(os.path.join(RESULTS_DIR, file)) as f:
                        entries.append(json.load(f))
                except Exception as e:
                    print(f"Error reading {file}: {e}")
    return {"results": entries}


@router.get("/{run_id}/step/{step_name}")
def get_step_details(run_id: str, step_name: str):
    """
    Get detailed information for a specific validation step.

    This endpoint returns comprehensive step details for any validation type,
    suitable for displaying in a drill-down modal.

    Args:
        run_id: The pipeline run ID
        step_name: The validation step name

    Returns:
        Step details including status, messages, error counts, SQL queries, and any available comparison data
    """
    # Find the result file for this run_id
    if not os.path.exists(RESULTS_DIR):
        raise HTTPException(status_code=404, detail="Results directory not found")

    result_file = None
    for file in os.listdir(RESULTS_DIR):
        if file.endswith(".json") and run_id in file:
            result_file = os.path.join(RESULTS_DIR, file)
            break

    if not result_file:
        raise HTTPException(status_code=404, detail=f"No results found for run_id: {run_id}")

    try:
        with open(result_file) as f:
            result_data = json.load(f)

        # Find the specific step (try both "steps" and "results" keys)
        steps = result_data.get("steps", result_data.get("results", []))
        target_step = None

        for step in steps:
            # Check both "step_name" and "name" keys
            step_id = step.get("step_name", step.get("name"))
            if step_id == step_name:
                target_step = step
                break

        if not target_step:
            raise HTTPException(status_code=404, detail=f"Step '{step_name}' not found in run {run_id}")

        # Extract details
        details = target_step.get("details", {})

        # Build response with all available information
        response = {
            "run_id": run_id,
            "step_name": step_name,
            "status": target_step.get("status", "unknown"),
            "severity": target_step.get("severity", details.get("severity")),
            "validation_type": target_step.get("validation_type", details.get("validation_type")),
            "message": target_step.get("message", details.get("message", "")),
            "error_count": target_step.get("error_count", details.get("error_count", 0)),
            "execution_time": target_step.get("execution_time", details.get("execution_time")),
        }

        # Add error details if available
        if target_step.get("errors") or details.get("errors"):
            response["errors"] = target_step.get("errors", details.get("errors"))

        # Add SQL queries if available
        if details.get("sql_query") or details.get("snow_query"):
            response["queries"] = {
                "sql_query": details.get("sql_query"),
                "snow_query": details.get("snow_query")
            }

        # Add comparison details for table comparison validations
        if "comparison_details" in target_step or "comparison_details" in details:
            comparison_details = target_step.get("comparison_details", details.get("comparison_details"))
            response["has_comparison_data"] = True
            response["comparison_summary"] = {
                "total_rows": target_step.get("sql_row_count", details.get("sql_row_count", 0)),
                "differing_rows": target_step.get("differing_rows_count", details.get("differing_rows_count", 0)),
                "affected_columns": target_step.get("affected_columns", details.get("affected_columns", [])),
                "difference_type": target_step.get("difference_type", details.get("difference_type"))
            }
        else:
            response["has_comparison_data"] = False

        # Add any additional metadata
        if details.get("metadata"):
            response["metadata"] = details["metadata"]

        return response

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving step details: {str(e)}")


@router.get("/{run_id}/step/{step_name}/comparison")
def get_comparison_details(run_id: str, step_name: str):
    """
    Get detailed comparison data for a specific validation step.

    This endpoint returns the full side-by-side comparison data that can be
    used to render a detailed comparison view in the UI.

    Args:
        run_id: The pipeline run ID
        step_name: The validation step name

    Returns:
        Comparison details including:
        - columns: List of column names
        - rows: Array of row comparisons with sql_values, snowflake_values, and differing_columns
        - summary: High-level statistics about the comparison
    """
    # Find the result file for this run_id
    if not os.path.exists(RESULTS_DIR):
        return {"error": "Results directory not found"}

    result_file = None
    for file in os.listdir(RESULTS_DIR):
        if file.endswith(".json") and run_id in file:
            result_file = os.path.join(RESULTS_DIR, file)
            break

    if not result_file:
        return {"error": f"No results found for run_id: {run_id}"}

    try:
        with open(result_file) as f:
            result_data = json.load(f)

        # Find the specific step (try both "steps" and "results" keys)
        steps = result_data.get("steps", result_data.get("results", []))
        target_step = None

        for step in steps:
            # Check both "step_name" and "name" keys
            step_id = step.get("step_name", step.get("name"))
            if step_id == step_name:
                target_step = step
                break

        if not target_step:
            return {"error": f"Step '{step_name}' not found in run {run_id}"}

        # Check if comparison_details exists (check both at top level and in details)
        details = target_step.get("details", {})
        if "comparison_details" not in target_step and "comparison_details" not in details:
            return {
                "error": "No comparison details available for this step",
                "message": target_step.get("message", details.get("message", "No detailed comparison data"))
            }

        # Return the comparison details with summary (check both locations)
        comparison_details = target_step.get("comparison_details", details.get("comparison_details"))

        return {
            "run_id": run_id,
            "step_name": step_name,
            "status": target_step.get("status"),
            "difference_type": target_step.get("difference_type", details.get("difference_type", "unknown")),
            "summary": {
                "total_rows": target_step.get("sql_row_count", details.get("sql_row_count", 0)),
                "differing_rows": target_step.get("differing_rows_count", details.get("differing_rows_count", 0)),
                "affected_columns": target_step.get("affected_columns", details.get("affected_columns", [])),
                "message": target_step.get("message", details.get("message", ""))
            },
            "comparison": comparison_details,
            "queries": {
                "sql_query": details.get("sql_query"),
                "snow_query": details.get("snow_query")
            } if details.get("sql_query") or details.get("snow_query") else None
        }

    except Exception as e:
        return {"error": f"Failed to load comparison details: {str(e)}"}


@router.get("/compare/{run_id_1}/vs/{run_id_2}")
def compare_pipeline_runs(run_id_1: str, run_id_2: str):
    """
    Compare two pipeline runs to show improvements and regressions.

    Returns a comprehensive comparison showing:
    - Step-by-step comparison
    - Error count changes (increased/decreased)
    - New errors introduced
    - Fixed errors
    - Overall trend (improving/degrading/stable)
    """
    try:
        # Load both run results
        run1_data = _load_run_data(run_id_1)
        run2_data = _load_run_data(run_id_2)

        if "error" in run1_data:
            raise HTTPException(status_code=404, detail=f"Run 1 not found: {run1_data['error']}")
        if "error" in run2_data:
            raise HTTPException(status_code=404, detail=f"Run 2 not found: {run2_data['error']}")

        # Extract steps from both runs
        steps1 = run1_data.get("steps", run1_data.get("results", []))
        steps2 = run2_data.get("steps", run2_data.get("results", []))

        # Create step lookup by name
        steps1_map = {step.get("step_name", step.get("name")): step for step in steps1}
        steps2_map = {step.get("step_name", step.get("name")): step for step in steps2}

        # Compare each step
        step_comparisons = []
        all_step_names = set(steps1_map.keys()) | set(steps2_map.keys())

        total_errors_run1 = 0
        total_errors_run2 = 0
        improved_steps = 0
        degraded_steps = 0
        stable_steps = 0
        new_errors = []
        fixed_errors = []

        for step_name in sorted(all_step_names):
            step1 = steps1_map.get(step_name)
            step2 = steps2_map.get(step_name)

            comparison = {
                "step_name": step_name,
                "exists_in_run1": step1 is not None,
                "exists_in_run2": step2 is not None
            }

            if step1 and step2:
                # Both steps exist - compare them
                status1 = step1.get("status", "unknown")
                status2 = step2.get("status", "unknown")

                # Count errors
                errors1 = _count_errors_in_step(step1)
                errors2 = _count_errors_in_step(step2)

                total_errors_run1 += errors1
                total_errors_run2 += errors2

                error_delta = errors2 - errors1

                comparison.update({
                    "run1_status": status1,
                    "run2_status": status2,
                    "run1_errors": errors1,
                    "run2_errors": errors2,
                    "error_delta": error_delta,
                    "status_changed": status1 != status2,
                    "trend": "improved" if error_delta < 0 else ("degraded" if error_delta > 0 else "stable")
                })

                if error_delta < 0:
                    improved_steps += 1
                elif error_delta > 0:
                    degraded_steps += 1
                else:
                    stable_steps += 1

            elif step1 and not step2:
                # Step removed in run2
                comparison["change"] = "removed"
                errors1 = _count_errors_in_step(step1)
                total_errors_run1 += errors1
                comparison["run1_errors"] = errors1

            elif not step1 and step2:
                # New step in run2
                comparison["change"] = "added"
                errors2 = _count_errors_in_step(step2)
                total_errors_run2 += errors2
                comparison["run2_errors"] = errors2
                new_errors.append(step_name)

            step_comparisons.append(comparison)

        # Overall trend
        overall_error_delta = total_errors_run2 - total_errors_run1
        if overall_error_delta < 0:
            overall_trend = "improving"
        elif overall_error_delta > 0:
            overall_trend = "degrading"
        else:
            overall_trend = "stable"

        # Calculate executive summary for run2 (comparison/newer run)
        executive_summary = _calculate_executive_summary(run2_data, steps2)

        # Calculate trend analysis
        pipeline_name = run2_data.get("batch_job_name") or run2_data.get("pipeline_name")
        trend_analysis = _calculate_trend_analysis(run2_data, pipeline_name)

        # Add severity to each step comparison
        for comparison in step_comparisons:
            step_name = comparison["step_name"]
            step2 = steps2_map.get(step_name)
            if step2:
                error_count = _count_errors_in_step(step2)
                if error_count > 0 or step2.get("status") in ["ERROR", "FAILED"]:
                    comparison["severity"] = _classify_severity(step2, error_count)

        # Group errors by root cause
        root_cause_groups = _group_by_root_cause(step_comparisons)

        # Generate actionable recommendations
        recommendations = _generate_actionable_recommendations(
            root_cause_groups=root_cause_groups,
            step_comparisons=step_comparisons,
            trend_analysis=trend_analysis,
            executive_summary=executive_summary
        )

        # Calculate financial impact and risk assessment
        financial_impact = _calculate_financial_impact(
            step_comparisons=step_comparisons,
            root_cause_groups=root_cause_groups,
            executive_summary=executive_summary
        )

        return {
            "status": "success",
            "run1": {
                "run_id": run_id_1,
                "timestamp": run1_data.get("timestamp", run1_data.get("execution_time")),
                "total_errors": total_errors_run1,
                "total_steps": len(steps1_map)
            },
            "run2": {
                "run_id": run_id_2,
                "timestamp": run2_data.get("timestamp", run2_data.get("execution_time")),
                "total_errors": total_errors_run2,
                "total_steps": len(steps2_map)
            },
            "comparison_summary": {
                "overall_trend": overall_trend,
                "total_error_delta": overall_error_delta,
                "error_delta_percentage": round((overall_error_delta / max(total_errors_run1, 1)) * 100, 2),
                "improved_steps": improved_steps,
                "degraded_steps": degraded_steps,
                "stable_steps": stable_steps,
                "new_steps": len([c for c in step_comparisons if c.get("change") == "added"]),
                "removed_steps": len([c for c in step_comparisons if c.get("change") == "removed"])
            },
            "executive_summary": executive_summary,
            "trend_analysis": trend_analysis,
            "root_cause_groups": root_cause_groups,
            "recommendations": recommendations,
            "financial_impact": financial_impact,
            "step_comparisons": step_comparisons
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Comparison failed: {str(e)}")


def _generate_recommendations(problematic_steps: list, trend_direction: str, health_score: float, error_trend_data: list) -> list:
    """Generate recommendations based on project health metrics"""
    recommendations = []

    # Recommendation based on health score
    if health_score < 50:
        recommendations.append({
            "priority": "CRITICAL",
            "title": "Critical Health Score",
            "description": f"Project health score is critically low at {health_score:.1f}%. Immediate action required.",
            "action": "Focus on resolving top 3 problematic validation steps immediately"
        })
    elif health_score < 70:
        recommendations.append({
            "priority": "HIGH",
            "title": "Low Health Score",
            "description": f"Project health score is below target at {health_score:.1f}%.",
            "action": "Allocate additional resources to address validation failures"
        })

    # Recommendation based on trend
    if trend_direction == "degrading":
        recommendations.append({
            "priority": "HIGH",
            "title": "Degrading Trend Detected",
            "description": "Validation errors are increasing over recent runs.",
            "action": "Review recent code changes and identify regression sources"
        })
    elif trend_direction == "stable" and health_score < 90:
        recommendations.append({
            "priority": "MEDIUM",
            "title": "Progress Stalled",
            "description": "Error rate has plateaued but hasn't reached target.",
            "action": "Consider different approach or additional testing strategies"
        })

    # Recommendation based on problematic steps
    if problematic_steps:
        top_step = problematic_steps[0]
        recommendations.append({
            "priority": "HIGH",
            "title": f"Address Top Failing Step: {top_step['step_name']}",
            "description": f"This step has failed {top_step['failure_count']} times ({top_step['failure_rate']:.1f}% failure rate).",
            "action": f"Investigate and fix validation logic for {top_step['step_name']}"
        })

    # Default recommendation if health is good
    if health_score >= 90 and trend_direction in ["improving", "stable"]:
        recommendations.append({
            "priority": "LOW",
            "title": "Maintain Current Quality",
            "description": f"Project health is excellent at {health_score:.1f}%.",
            "action": "Continue current practices and monitor for any regressions"
        })

    return recommendations


@router.get("/project-summary")
def get_project_summary():
    """
    Generate a comprehensive project summary for Tech Lead view.

    Analyzes all pipeline runs to provide:
    - Error trends over time
    - Most problematic validation steps
    - Success rate trends
    - Overall project health score
    - Recommendations
    """
    try:
        # Load all results
        all_results = []
        if os.path.exists(RESULTS_DIR):
            for file in os.listdir(RESULTS_DIR):
                if file.endswith(".json"):
                    try:
                        with open(os.path.join(RESULTS_DIR, file)) as f:
                            all_results.append(json.load(f))
                    except Exception as e:
                        print(f"Error reading {file}: {e}")

        if not all_results:
            return {
                "status": "no_data",
                "message": "No pipeline runs found"
            }

        # Sort by timestamp
        all_results.sort(key=lambda x: x.get("timestamp", x.get("execution_time", "")))

        # Analyze trends
        error_trend_data = []
        step_error_counts = defaultdict(int)
        step_failure_counts = defaultdict(int)
        step_total_runs = defaultdict(int)
        total_runs = len(all_results)
        total_errors_all_runs = 0
        total_steps_all_runs = 0
        total_failed_steps = 0

        # Pipeline breakdown: separate standard vs workload validations
        workload_runs = []
        standard_runs = []
        workload_errors = 0
        workload_steps = 0
        workload_failed_steps = 0
        standard_errors = 0
        standard_steps = 0
        standard_failed_steps = 0

        for run_index, result in enumerate(all_results, 1):
            steps = result.get("steps", result.get("results", []))
            run_errors = 0
            run_failed_steps = 0

            # Determine if this is a workload-based pipeline
            is_workload = False
            execution_type = result.get("execution_type", "")
            batch_job_name = result.get("batch_job_name", "")

            # Check if any validation name contains "workload_query"
            for step in steps:
                step_name = step.get("step_name", step.get("name", ""))
                if "workload_query" in step_name.lower():
                    is_workload = True
                    break

            # Also check for batch consolidation type
            if execution_type == "batch_consolidation" or "batch" in batch_job_name.lower():
                is_workload = True

            for step in steps:
                step_name = step.get("step_name", step.get("name"))
                step_total_runs[step_name] += 1

                errors = _count_errors_in_step(step)
                run_errors += errors

                if errors > 0:
                    step_error_counts[step_name] += errors
                    step_failure_counts[step_name] += 1
                    run_failed_steps += 1

            total_errors_all_runs += run_errors
            total_steps_all_runs += len(steps)
            total_failed_steps += run_failed_steps

            # Categorize into workload or standard
            if is_workload:
                workload_runs.append(result.get("run_id"))
                workload_errors += run_errors
                workload_steps += len(steps)
                workload_failed_steps += run_failed_steps
            else:
                standard_runs.append(result.get("run_id"))
                standard_errors += run_errors
                standard_steps += len(steps)
                standard_failed_steps += run_failed_steps

            error_trend_data.append({
                "run_id": result.get("run_id"),
                "run_number": run_index,
                "timestamp": result.get("timestamp") or result.get("started_at") or result.get("execution_time"),
                "total_errors": run_errors,
                "total_steps": len(steps),
                "failed_steps": run_failed_steps,
                "pipeline_type": "workload" if is_workload else "standard"
            })

        # Calculate most problematic steps
        problematic_steps = []
        for step_name in step_error_counts.keys():
            failure_rate = (step_failure_counts[step_name] / step_total_runs[step_name]) * 100
            problematic_steps.append({
                "step_name": step_name,
                "total_errors": step_error_counts[step_name],
                "failure_count": step_failure_counts[step_name],
                "total_runs": step_total_runs[step_name],
                "failure_rate": round(failure_rate, 2)
            })

        # Sort by total errors descending
        problematic_steps.sort(key=lambda x: x["total_errors"], reverse=True)

        # Calculate trend (last 5 runs vs previous runs)
        recent_runs = error_trend_data[-5:] if len(error_trend_data) >= 5 else error_trend_data
        older_runs = error_trend_data[:-5] if len(error_trend_data) > 5 else []

        recent_avg_errors = sum(r["total_errors"] for r in recent_runs) / len(recent_runs) if recent_runs else 0
        older_avg_errors = sum(r["total_errors"] for r in older_runs) / len(older_runs) if older_runs else recent_avg_errors

        if recent_avg_errors < older_avg_errors * 0.9:
            trend_direction = "improving"
        elif recent_avg_errors > older_avg_errors * 1.1:
            trend_direction = "degrading"
        else:
            trend_direction = "stable"

        # Calculate health score (0-100) based on step failure rate
        # This accounts for ALL validation types: standard validations, custom SQL, workload analysis, etc.
        if total_steps_all_runs > 0:
            failure_rate = (total_failed_steps / total_steps_all_runs) * 100
            health_score = max(0, 100 - failure_rate)
        else:
            health_score = 100.0

        # Calculate average errors per run (for display purposes)
        avg_errors_per_run = total_errors_all_runs / total_runs if total_runs > 0 else 0

        # Generate recommendations
        recommendations = _generate_recommendations(
            problematic_steps,
            trend_direction,
            health_score,
            error_trend_data
        )

        # Calculate pass rates for breakdown
        workload_pass_rate = 0.0
        if workload_steps > 0:
            workload_pass_rate = ((workload_steps - workload_failed_steps) / workload_steps) * 100

        standard_pass_rate = 0.0
        if standard_steps > 0:
            standard_pass_rate = ((standard_steps - standard_failed_steps) / standard_steps) * 100

        # Calculate average errors per run for each type
        workload_avg_errors = workload_errors / len(workload_runs) if workload_runs else 0
        standard_avg_errors = standard_errors / len(standard_runs) if standard_runs else 0

        return {
            "status": "success",
            "summary": {
                "total_runs": total_runs,
                "total_errors_all_time": total_errors_all_runs,
                "average_errors_per_run": round(avg_errors_per_run, 2),
                "health_score": round(health_score, 2),
                "trend_direction": trend_direction,
                "recent_avg_errors": round(recent_avg_errors, 2),
                "older_avg_errors": round(older_avg_errors, 2)
            },
            "pipeline_breakdown": {
                "standard": {
                    "total_runs": len(standard_runs),
                    "total_steps": standard_steps,
                    "failed_steps": standard_failed_steps,
                    "total_errors": standard_errors,
                    "pass_rate": round(standard_pass_rate, 2),
                    "avg_errors_per_run": round(standard_avg_errors, 2)
                },
                "workload": {
                    "total_runs": len(workload_runs),
                    "total_steps": workload_steps,
                    "failed_steps": workload_failed_steps,
                    "total_errors": workload_errors,
                    "pass_rate": round(workload_pass_rate, 2),
                    "avg_errors_per_run": round(workload_avg_errors, 2)
                }
            },
            "error_trend": error_trend_data,
            "problematic_steps": problematic_steps[:10],  # Top 10
            "recommendations": recommendations,
            "latest_run": error_trend_data[-1] if error_trend_data else None,
            "oldest_run": error_trend_data[0] if error_trend_data else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate project summary: {str(e)}")


def _load_run_data(run_id: str) -> Dict:
    """Load run data from results directory"""
    if not os.path.exists(RESULTS_DIR):
        return {"error": "Results directory not found"}

    for file in os.listdir(RESULTS_DIR):
        if file.endswith(".json") and run_id in file:
            try:
                with open(os.path.join(RESULTS_DIR, file)) as f:
                    return json.load(f)
            except Exception as e:
                return {"error": f"Failed to load run data: {str(e)}"}

    return {"error": f"Run {run_id} not found"}


def _count_errors_in_step(step: Dict) -> int:
    """Count total errors in a validation step"""
    # Try different possible error count keys
    error_count = 0

    # Direct error count
    if "error_count" in step:
        error_count = step["error_count"]
    elif "differing_rows_count" in step:
        error_count = step["differing_rows_count"]

    # Check details
    details = step.get("details", {})
    if "differing_rows_count" in details:
        error_count = max(error_count, details["differing_rows_count"])

    # Check if step failed or has ERROR status
    if (step.get("status") in ["failed", "ERROR"]) and error_count == 0:
        error_count = 1  # At least count as 1 error if failed

    return error_count


def _classify_severity(step: Dict, error_count: int) -> str:
    """
    Classify severity of validation failures.

    Severity Rules:
    - BLOCKER: Row count diff >5%, schema mismatch, FK violations, ERROR status
    - HIGH: Null values in critical cols, data type mismatches, >1000 errors
    - MEDIUM: Performance issues, index differences, 100-1000 errors
    - LOW: Formatting, whitespace, case differences, <100 errors
    """
    status = step.get("status", "").upper()
    step_name = step.get("step_name", step.get("name", "")).lower()
    details = step.get("details", {})

    # BLOCKER conditions
    if status == "ERROR":
        return "BLOCKER"

    if "schema" in step_name or "foreign_key" in step_name or "fk" in step_name:
        return "BLOCKER"

    # Check for row count differences
    if "row_count" in step_name or "record_count" in step_name:
        if error_count > 0:
            return "BLOCKER"

    # Check error message for critical patterns
    error_msg = details.get("error", "").lower()
    if any(pattern in error_msg for pattern in ["parameter mismatch", "schema", "foreign key"]):
        return "BLOCKER"

    # HIGH conditions
    if error_count > 1000:
        return "HIGH"

    if "null" in step_name or "datatype" in step_name or "data_type" in step_name:
        return "HIGH"

    if step.get("severity") == "HIGH":  # Respect existing severity if present
        return "HIGH"

    # MEDIUM conditions
    if 100 <= error_count <= 1000:
        return "MEDIUM"

    if "performance" in step_name or "index" in step_name:
        return "MEDIUM"

    # LOW conditions (default for small error counts)
    return "LOW"


def _calculate_executive_summary(run_data: Dict, steps: List[Dict]) -> Dict:
    """
    Calculate executive summary metrics for a pipeline run.

    Returns:
    - readiness_score: 0-100 score based on validation pass rate
    - overall_status: "Ready" (>90%), "On Track" (70-90%), "At Risk" (<70%)
    - total_validations: Total number of validation steps
    - passed_validations: Number of successful validations
    - warnings: Number of MEDIUM severity issues
    - critical_issues: Number of BLOCKER/HIGH severity issues
    - severity_breakdown: Count of each severity level
    """
    total_validations = len(steps)
    passed_validations = 0
    severity_counts = {"BLOCKER": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}

    for step in steps:
        status = step.get("status", "").upper()
        error_count = _count_errors_in_step(step)

        # Count passed validations
        if status in ["SUCCESS", "PASS"] or (status not in ["ERROR", "FAILED"] and error_count == 0):
            passed_validations += 1

        # Classify severity
        if status in ["ERROR", "FAILED"] or error_count > 0:
            severity = _classify_severity(step, error_count)
            severity_counts[severity] += 1

    # Calculate readiness score
    if total_validations > 0:
        readiness_score = (passed_validations / total_validations) * 100
    else:
        readiness_score = 0

    # Determine overall status
    if readiness_score >= 90:
        overall_status = "Ready"
    elif readiness_score >= 70:
        overall_status = "On Track"
    else:
        overall_status = "At Risk"

    # Critical issues = BLOCKER + HIGH
    critical_issues = severity_counts["BLOCKER"] + severity_counts["HIGH"]
    warnings = severity_counts["MEDIUM"]

    return {
        "readiness_score": round(readiness_score, 2),
        "overall_status": overall_status,
        "total_validations": total_validations,
        "passed_validations": passed_validations,
        "failed_validations": total_validations - passed_validations,
        "warnings": warnings,
        "critical_issues": critical_issues,
        "severity_breakdown": severity_counts
    }


def _calculate_trend_analysis(run2_data: Dict, pipeline_name: str = None) -> Dict:
    """
    Calculate trend analysis from historical runs.

    Returns:
    - error_trend: List of historical error counts over time
    - velocity: Error reduction rate per week
    - projected_zero_date: Estimated date when errors will reach zero
    - regression_detected: Boolean indicating if errors are increasing
    """
    try:
        # Load all historical runs for the same pipeline/batch
        historical_runs = []

        if not os.path.exists(RESULTS_DIR):
            return _empty_trend_analysis()

        # Determine what we're tracking (batch or pipeline)
        target_name = run2_data.get("batch_job_name") or run2_data.get("pipeline_name") or pipeline_name
        if not target_name:
            return _empty_trend_analysis()

        # Load all matching runs
        for file in sorted(os.listdir(RESULTS_DIR)):
            if not file.endswith(".json"):
                continue

            file_path = os.path.join(RESULTS_DIR, file)
            try:
                with open(file_path, 'r') as f:
                    run_data = json.load(f)

                # Match by batch name or pipeline name
                run_name = run_data.get("batch_job_name") or run_data.get("pipeline_name")
                if run_name == target_name:
                    # Extract timestamp
                    timestamp_str = run_data.get("started_at") or run_data.get("timestamp") or run_data.get("execution_time")
                    if not timestamp_str:
                        continue

                    timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))

                    # Count total errors
                    steps = run_data.get("steps", run_data.get("results", []))
                    total_errors = sum(_count_errors_in_step(step) for step in steps)
                    total_validations = len(steps)

                    historical_runs.append({
                        "run_id": run_data.get("run_id"),
                        "timestamp": timestamp,
                        "total_errors": total_errors,
                        "total_validations": total_validations,
                        "error_rate": (total_errors / max(total_validations, 1)) * 100
                    })
            except Exception as e:
                print(f"[TREND] Error processing {file}: {e}")
                continue

        # Sort by timestamp
        historical_runs.sort(key=lambda x: x["timestamp"])

        # Keep last 10 runs for trend analysis
        recent_runs = historical_runs[-10:] if len(historical_runs) > 10 else historical_runs

        if len(recent_runs) < 2:
            return _empty_trend_analysis()

        # Calculate velocity (error reduction per day)
        first_run = recent_runs[0]
        last_run = recent_runs[-1]

        time_diff_days = (last_run["timestamp"] - first_run["timestamp"]).total_days
        if time_diff_days == 0:
            time_diff_days = 1  # Avoid division by zero

        error_diff = last_run["total_errors"] - first_run["total_errors"]
        velocity_per_day = error_diff / time_diff_days
        velocity_per_week = velocity_per_day * 7

        # Detect regression (errors increasing)
        regression_detected = error_diff > 0

        # Project zero-error date
        projected_zero_date = None
        if velocity_per_day < 0 and last_run["total_errors"] > 0:
            # Negative velocity means errors decreasing
            days_to_zero = abs(last_run["total_errors"] / velocity_per_day)
            projected_zero_date = (last_run["timestamp"] + timedelta(days=days_to_zero)).isoformat()

        # Calculate percentage change
        if first_run["total_errors"] > 0:
            percentage_change = ((last_run["total_errors"] - first_run["total_errors"]) / first_run["total_errors"]) * 100
        else:
            percentage_change = 0

        return {
            "error_trend": [
                {
                    "timestamp": run["timestamp"].isoformat(),
                    "total_errors": run["total_errors"],
                    "error_rate": run["error_rate"]
                }
                for run in recent_runs
            ],
            "velocity": {
                "per_day": round(velocity_per_day, 2),
                "per_week": round(velocity_per_week, 2),
                "percentage_change": round(percentage_change, 2)
            },
            "projected_zero_date": projected_zero_date,
            "regression_detected": regression_detected,
            "total_runs_analyzed": len(recent_runs),
            "time_span_days": round(time_diff_days, 1)
        }

    except Exception as e:
        print(f"[TREND] Error calculating trend analysis: {e}")
        return _empty_trend_analysis()


def _empty_trend_analysis() -> Dict:
    """Return empty trend analysis structure"""
    return {
        "error_trend": [],
        "velocity": {
            "per_day": 0,
            "per_week": 0,
            "percentage_change": 0
        },
        "projected_zero_date": None,
        "regression_detected": False,
        "total_runs_analyzed": 0,
        "time_span_days": 0
    }


def _group_by_root_cause(step_comparisons: List[Dict]) -> List[Dict]:
    """
    Group validation failures by root cause patterns.

    Analyzes step comparison results and groups them by common error patterns
    such as schema mismatches, FK violations, null handling, etc.

    Returns:
        List of root cause groups with affected steps and recommended actions
    """
    try:
        root_causes = {}

        # Analyze each step comparison
        for step in step_comparisons:
            if not step.get("exists_in_run1") or not step.get("exists_in_run2"):
                continue

            step_name = step.get("step_name", "")
            severity = step.get("severity", "LOW")
            errors = step.get("run2_errors", 0)

            # Skip steps with no errors
            if errors == 0:
                continue

            # Categorize by pattern matching
            categories = []

            # Schema-related issues
            if any(keyword in step_name.lower() for keyword in ["schema", "datatype", "column"]):
                categories.append("schema_mismatch")

            # Foreign key violations
            if any(keyword in step_name.lower() for keyword in ["foreign_key", "fk", "referential"]):
                categories.append("foreign_key_violation")

            # Null handling issues
            if any(keyword in step_name.lower() for keyword in ["null", "nullability"]):
                categories.append("null_handling")

            # Row count discrepancies
            if any(keyword in step_name.lower() for keyword in ["record_count", "row_count", "count"]):
                categories.append("row_count_mismatch")

            # Date/timestamp issues
            if any(keyword in step_name.lower() for keyword in ["date", "timestamp", "time"]):
                categories.append("date_format")

            # Data quality issues
            if any(keyword in step_name.lower() for keyword in ["distribution", "outlier", "uniqueness", "pattern"]):
                categories.append("data_quality")

            # Dimension validation issues
            if any(keyword in step_name.lower() for keyword in ["dim_", "scd", "surrogate", "business_key"]):
                categories.append("dimension_validation")

            # Fact validation issues
            if any(keyword in step_name.lower() for keyword in ["fact_", "metric", "measure"]):
                categories.append("fact_validation")

            # If no specific category, categorize by severity
            if not categories:
                if severity == "BLOCKER":
                    categories.append("critical_blocker")
                elif severity == "HIGH":
                    categories.append("high_priority")
                else:
                    categories.append("other_issues")

            # Add step to each matching category
            for category in categories:
                if category not in root_causes:
                    root_causes[category] = {
                        "steps": [],
                        "total_errors": 0,
                        "max_severity": "LOW"
                    }

                root_causes[category]["steps"].append(step_name)
                root_causes[category]["total_errors"] += errors

                # Update max severity
                severity_order = {"LOW": 0, "MEDIUM": 1, "HIGH": 2, "BLOCKER": 3}
                current_max = root_causes[category]["max_severity"]
                if severity_order.get(severity, 0) > severity_order.get(current_max, 0):
                    root_causes[category]["max_severity"] = severity

        # Convert to list format with metadata
        category_metadata = {
            "schema_mismatch": {
                "title": "Schema Mismatches",
                "description": "Column name differences, data type mismatches, or structural changes between source and target",
                "recommendation": "Review schema mapping configuration and ensure data type compatibility"
            },
            "foreign_key_violation": {
                "title": "Foreign Key Violations",
                "description": "Referential integrity issues where child records reference non-existent parent records",
                "recommendation": "Ensure parent tables are loaded before child tables and verify referential integrity constraints"
            },
            "null_handling": {
                "title": "Null Value Issues",
                "description": "Unexpected null values in non-nullable columns or missing required data",
                "recommendation": "Review null handling logic and ensure default values are properly configured"
            },
            "row_count_mismatch": {
                "title": "Row Count Discrepancies",
                "description": "Significant differences in record counts between source and target systems",
                "recommendation": "Investigate data filtering logic and verify all source records are being migrated"
            },
            "date_format": {
                "title": "Date/Timestamp Format Issues",
                "description": "Date format mismatches or timezone conversion problems",
                "recommendation": "Standardize date formats and ensure timezone handling is consistent"
            },
            "data_quality": {
                "title": "Data Quality Issues",
                "description": "Distribution anomalies, outliers, uniqueness violations, or pattern mismatches",
                "recommendation": "Review data cleansing rules and validate data quality metrics"
            },
            "dimension_validation": {
                "title": "Dimension Validation Failures",
                "description": "Issues with dimension tables including SCD logic, surrogate keys, or business key mismatches",
                "recommendation": "Verify SCD implementation and ensure proper dimension key management"
            },
            "fact_validation": {
                "title": "Fact Table Validation Failures",
                "description": "Metric calculations, measure aggregations, or fact-dimension conformance issues",
                "recommendation": "Review fact table loading logic and verify dimensional relationships"
            },
            "critical_blocker": {
                "title": "Critical Blockers",
                "description": "Severe issues that must be resolved before migration can proceed",
                "recommendation": "Prioritize resolution of these issues immediately - migration cannot proceed until fixed"
            },
            "high_priority": {
                "title": "High Priority Issues",
                "description": "Important issues that should be addressed soon",
                "recommendation": "Schedule resolution of these issues in the next sprint"
            },
            "other_issues": {
                "title": "Other Validation Issues",
                "description": "Miscellaneous validation failures that need investigation",
                "recommendation": "Review and categorize these issues for appropriate handling"
            }
        }

        result = []
        for category, data in root_causes.items():
            metadata = category_metadata.get(category, {
                "title": category.replace("_", " ").title(),
                "description": f"Issues related to {category}",
                "recommendation": "Review and address these validation failures"
            })

            result.append({
                "category": category,
                "title": metadata["title"],
                "description": metadata["description"],
                "affected_steps": data["steps"],
                "total_affected": len(data["steps"]),
                "total_errors": data["total_errors"],
                "severity": data["max_severity"],
                "recommended_action": metadata["recommendation"]
            })

        # Sort by severity and total errors
        severity_order = {"BLOCKER": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1}
        result.sort(key=lambda x: (severity_order.get(x["severity"], 0), x["total_errors"]), reverse=True)

        return result

    except Exception as e:
        print(f"[ROOT_CAUSE] Error grouping by root cause: {e}")
        return []


def _generate_actionable_recommendations(
    root_cause_groups: List[Dict],
    step_comparisons: List[Dict],
    trend_analysis: Dict,
    executive_summary: Dict
) -> List[Dict]:
    """
    Generate actionable recommendations with priority classification and effort estimation.

    Returns list of recommendations with:
    - priority: P1 (Critical), P2 (High), P3 (Medium)
    - title: Short recommendation title
    - description: Detailed explanation
    - action_items: List of specific actions to take
    - commands: Copy-pasteable commands (if applicable)
    - effort: Estimated effort (Low/Medium/High)
    - impact: Expected impact (Low/Medium/High)
    """
    try:
        recommendations = []

        # Priority 1: Critical blockers from root cause analysis
        blockers = [g for g in root_cause_groups if g.get("severity") == "BLOCKER"]
        for blocker in blockers:
            rec = {
                "priority": "P1",
                "title": f"Critical: {blocker['title']}",
                "description": blocker["description"],
                "action_items": [
                    blocker["recommended_action"],
                    f"Review {blocker['total_affected']} affected validation steps",
                    "Fix root cause before proceeding with migration"
                ],
                "commands": [],
                "effort": "High" if blocker["total_affected"] > 5 else "Medium",
                "impact": "High",
                "affected_count": blocker["total_affected"],
                "category": blocker["category"]
            }

            # Add specific commands based on category
            if "schema" in blocker["category"]:
                rec["commands"] = [
                    "# Review schema differences",
                    "SELECT COLUMN_NAME, DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'your_table'",
                    "# Consider adding migration script to align schemas"
                ]
            elif "foreign_key" in blocker["category"]:
                rec["commands"] = [
                    "# Check FK violations",
                    "SELECT * FROM target_table WHERE fk_column NOT IN (SELECT pk FROM parent_table)",
                    "# Fix orphaned records or add missing parent records"
                ]
            elif "null" in blocker["category"]:
                rec["commands"] = [
                    "# Identify null records",
                    "SELECT * FROM table WHERE critical_column IS NULL",
                    "# Add default values or fix ETL to populate nulls"
                ]

            recommendations.append(rec)

        # Priority 2: Regression detection
        if trend_analysis.get("regression_detected", False):
            velocity = trend_analysis.get("velocity", {})
            error_trend = trend_analysis.get("error_trend", [])

            if error_trend and len(error_trend) >= 2:
                recent_increase = error_trend[-1]["total_errors"] - error_trend[-2]["total_errors"]
                rec = {
                    "priority": "P1",
                    "title": "Regression Detected: Error Count Increasing",
                    "description": f"Recent runs show an increase of {recent_increase} errors. Velocity: {velocity.get('per_day', 0):.2f} errors/day.",
                    "action_items": [
                        "Review recent code changes or data updates",
                        "Compare last 2-3 pipeline runs to identify new failures",
                        "Roll back changes if regression is severe",
                        "Add tests to prevent future regressions"
                    ],
                    "commands": [
                        "# Compare recent runs",
                        f"# Latest run: {error_trend[-1]['total_errors']} errors",
                        f"# Previous run: {error_trend[-2]['total_errors']} errors",
                        "# Review git commits between runs"
                    ],
                    "effort": "Medium",
                    "impact": "High",
                    "affected_count": recent_increase,
                    "category": "regression"
                }
                recommendations.append(rec)

        # Priority 2: High-severity issues
        high_priority = [g for g in root_cause_groups if g.get("severity") == "HIGH"]
        for high_issue in high_priority[:3]:  # Limit to top 3
            rec = {
                "priority": "P2",
                "title": f"High Priority: {high_issue['title']}",
                "description": high_issue["description"],
                "action_items": [
                    high_issue["recommended_action"],
                    f"Address {high_issue['total_errors']} errors across {high_issue['total_affected']} steps"
                ],
                "commands": [],
                "effort": "Medium",
                "impact": "High",
                "affected_count": high_issue["total_affected"],
                "category": high_issue["category"]
            }

            if "data_quality" in high_issue["category"]:
                rec["commands"] = [
                    "# Profile data quality",
                    "SELECT column, COUNT(*) as errors FROM validation_results WHERE status = 'FAILED' GROUP BY column",
                    "# Add data quality rules to ETL pipeline"
                ]

            recommendations.append(rec)

        # Priority 2: Low readiness score
        readiness_score = executive_summary.get("readiness_score", 100)
        if readiness_score < 70:
            status = executive_summary.get("overall_status", "Unknown")
            rec = {
                "priority": "P1" if readiness_score < 50 else "P2",
                "title": f"Low Readiness Score: {readiness_score:.1f}%",
                "description": f"Migration readiness is {status}. Need to achieve >90% for production readiness.",
                "action_items": [
                    "Focus on resolving BLOCKER and HIGH severity issues first",
                    "Target critical tables/dimensions before proceeding to facts",
                    "Allocate additional resources if timeline is tight",
                    f"Current gap: {90 - readiness_score:.1f} percentage points to target"
                ],
                "commands": [
                    "# Prioritize critical validations",
                    "# 1. Schema validations (structural alignment)",
                    "# 2. Foreign key validations (referential integrity)",
                    "# 3. Data quality validations (completeness, accuracy)"
                ],
                "effort": "High",
                "impact": "High",
                "affected_count": executive_summary.get("critical_issues", 0),
                "category": "readiness"
            }
            recommendations.append(rec)

        # Priority 3: Optimization opportunities
        if trend_analysis.get("projected_zero_date"):
            rec = {
                "priority": "P3",
                "title": "Projected Timeline to Zero Errors",
                "description": f"Based on current velocity, zero-error completion projected for {trend_analysis['projected_zero_date']}",
                "action_items": [
                    "Continue current pace of error resolution",
                    "Monitor velocity to detect slowdowns",
                    "Allocate more resources if timeline needs acceleration"
                ],
                "commands": [],
                "effort": "Low",
                "impact": "Medium",
                "affected_count": 0,
                "category": "timeline"
            }
            recommendations.append(rec)

        # Priority 3: Medium severity issues (top 2)
        medium_issues = [g for g in root_cause_groups if g.get("severity") == "MEDIUM"]
        for med_issue in medium_issues[:2]:
            rec = {
                "priority": "P3",
                "title": f"Medium Priority: {med_issue['title']}",
                "description": med_issue["description"],
                "action_items": [
                    med_issue["recommended_action"],
                    "Address after P1 and P2 items are resolved"
                ],
                "commands": [],
                "effort": "Low" if med_issue["total_affected"] < 3 else "Medium",
                "impact": "Medium",
                "affected_count": med_issue["total_affected"],
                "category": med_issue["category"]
            }
            recommendations.append(rec)

        # Priority 3: Success/Encouragement if doing well
        if readiness_score >= 90 and not trend_analysis.get("regression_detected"):
            rec = {
                "priority": "P3",
                "title": "Excellent Progress: Ready for Migration",
                "description": f"Readiness score of {readiness_score:.1f}% indicates strong migration readiness.",
                "action_items": [
                    "Review remaining warnings to ensure no hidden issues",
                    "Perform final UAT/testing in staging environment",
                    "Document any known limitations or workarounds",
                    "Schedule production migration cutover"
                ],
                "commands": [
                    "# Final validation checklist",
                    "# 1. All BLOCKER issues resolved",
                    "# 2. All HIGH issues resolved or documented exceptions",
                    "# 3. Stakeholder sign-off obtained",
                    "# 4. Rollback plan documented"
                ],
                "effort": "Low",
                "impact": "High",
                "affected_count": 0,
                "category": "success"
            }
            recommendations.append(rec)

        # Sort by priority (P1 > P2 > P3)
        priority_order = {"P1": 3, "P2": 2, "P3": 1}
        recommendations.sort(key=lambda x: priority_order.get(x["priority"], 0), reverse=True)

        return recommendations

    except Exception as e:
        print(f"[RECOMMENDATIONS] Error generating recommendations: {e}")
        return []


def _calculate_financial_impact(
    step_comparisons: List[Dict],
    root_cause_groups: List[Dict],
    executive_summary: Dict
) -> Dict:
    """
    Calculate financial impact and risk assessment for data quality issues.

    Returns:
    - table_criticality: List of tables with criticality scores
    - financial_impact: Estimated cost impact of issues
    - risk_assessment: Overall risk level and breakdown
    """
    try:
        # Table criticality scoring (0-10 scale)
        table_criticality_scores = {}

        for step in step_comparisons:
            step_name = step.get("step_name", "").lower()

            # Determine criticality based on table type and validation type
            criticality = 5  # Default medium criticality

            # FACT tables = highest criticality (8-10)
            if "fact" in step_name or "sales" in step_name or "transaction" in step_name:
                criticality = 9
            # DIM tables = high criticality (6-8)
            elif "dim" in step_name or "dimension" in step_name:
                criticality = 7
            # Schema validations = high criticality
            elif "schema" in step_name:
                criticality = 8
            # Foreign key validations = high criticality
            elif "foreign_key" in step_name or "fk" in step_name:
                criticality = 7
            # Data quality checks = medium-high
            elif any(x in step_name for x in ["null", "uniqueness", "distribution"]):
                criticality = 6
            # Performance/metrics = medium
            elif "metric" in step_name or "ratio" in step_name:
                criticality = 5

            # Adjust based on error severity
            if step.get("severity") == "BLOCKER":
                criticality = min(10, criticality + 2)
            elif step.get("severity") == "HIGH":
                criticality = min(10, criticality + 1)

            table_criticality_scores[step_name] = criticality

        # Build table criticality list
        table_criticality = []
        for step_name, score in sorted(table_criticality_scores.items(), key=lambda x: x[1], reverse=True):
            step = next((s for s in step_comparisons if s.get("step_name") == step_name), {})

            table_criticality.append({
                "table_name": step_name,
                "criticality_score": score,
                "criticality_level": "Critical" if score >= 8 else "High" if score >= 6 else "Medium" if score >= 4 else "Low",
                "error_count": step.get("run2_errors", 0),
                "status": step.get("run2_status", "UNKNOWN"),
                "severity": step.get("severity", "UNKNOWN")
            })

        # Financial impact calculation
        # Cost per error based on criticality (estimated in dollars)
        cost_per_error = {
            10: 5000,  # Critical: $5,000 per error
            9: 3000,   # Very High: $3,000 per error
            8: 2000,   # High: $2,000 per error
            7: 1000,   # Medium-High: $1,000 per error
            6: 500,    # Medium: $500 per error
            5: 250,    # Medium-Low: $250 per error
            4: 100,    # Low: $100 per error
            3: 50,     # Very Low: $50 per error
            2: 25,     # Minimal: $25 per error
            1: 10      # Negligible: $10 per error
        }

        total_estimated_cost = 0
        cost_breakdown = []

        for item in table_criticality:
            if item["error_count"] > 0:
                unit_cost = cost_per_error.get(item["criticality_score"], 100)
                item_cost = item["error_count"] * unit_cost
                total_estimated_cost += item_cost

                cost_breakdown.append({
                    "table_name": item["table_name"],
                    "error_count": item["error_count"],
                    "criticality_score": item["criticality_score"],
                    "unit_cost": unit_cost,
                    "total_cost": item_cost
                })

        # Risk assessment
        blocker_count = sum(1 for g in root_cause_groups if g.get("severity") == "BLOCKER")
        high_count = sum(1 for g in root_cause_groups if g.get("severity") == "HIGH")
        total_errors = executive_summary.get("failed_validations", 0)
        readiness_score = executive_summary.get("readiness_score", 100)

        # Determine overall risk level
        if blocker_count > 0 or readiness_score < 50:
            overall_risk = "Critical"
            risk_score = 9
        elif high_count > 3 or readiness_score < 70:
            overall_risk = "High"
            risk_score = 7
        elif total_errors > 10 or readiness_score < 85:
            overall_risk = "Medium"
            risk_score = 5
        else:
            overall_risk = "Low"
            risk_score = 3

        # Risk factors
        risk_factors = []

        if blocker_count > 0:
            risk_factors.append({
                "factor": "Blocker Issues Present",
                "impact": "Critical",
                "description": f"{blocker_count} blocker-level issues detected that prevent migration"
            })

        if readiness_score < 70:
            risk_factors.append({
                "factor": "Low Readiness Score",
                "impact": "High",
                "description": f"Readiness score of {readiness_score:.1f}% is below target threshold of 90%"
            })

        if total_estimated_cost > 100000:
            risk_factors.append({
                "factor": "High Financial Impact",
                "impact": "High",
                "description": f"Estimated cost impact of ${total_estimated_cost:,.0f} exceeds $100,000"
            })

        # Critical tables at risk
        critical_tables_at_risk = [
            t for t in table_criticality
            if t["criticality_score"] >= 8 and t["error_count"] > 0
        ]

        if critical_tables_at_risk:
            risk_factors.append({
                "factor": "Critical Tables Affected",
                "impact": "High",
                "description": f"{len(critical_tables_at_risk)} critical tables have validation errors"
            })

        return {
            "table_criticality": table_criticality[:10],  # Top 10 critical tables
            "financial_impact": {
                "total_estimated_cost": total_estimated_cost,
                "cost_breakdown": sorted(cost_breakdown, key=lambda x: x["total_cost"], reverse=True)[:10],
                "average_cost_per_error": total_estimated_cost / max(total_errors, 1),
                "high_cost_tables": len([c for c in cost_breakdown if c["total_cost"] > 10000])
            },
            "risk_assessment": {
                "overall_risk": overall_risk,
                "risk_score": risk_score,
                "risk_factors": risk_factors,
                "blocker_issues": blocker_count,
                "high_severity_issues": high_count,
                "critical_tables_at_risk": len(critical_tables_at_risk),
                "migration_readiness": "Not Ready" if risk_score >= 7 else "Ready with Conditions" if risk_score >= 5 else "Ready"
            }
        }

    except Exception as e:
        print(f"[FINANCIAL_IMPACT] Error calculating financial impact: {e}")
        return {
            "table_criticality": [],
            "financial_impact": {
                "total_estimated_cost": 0,
                "cost_breakdown": [],
                "average_cost_per_error": 0,
                "high_cost_tables": 0
            },
            "risk_assessment": {
                "overall_risk": "Unknown",
                "risk_score": 0,
                "risk_factors": [],
                "blocker_issues": 0,
                "high_severity_issues": 0,
                "critical_tables_at_risk": 0,
                "migration_readiness": "Unknown"
            }
        }

@router.get("/history")
def get_historical_trends(pipeline_name: str = None, limit: int = 10):
    """
    Get historical trend data for pipeline runs over time.
    
    Args:
        pipeline_name: Optional filter by pipeline name
        limit: Maximum number of runs to include (default 10)
    
    Returns:
        Historical metrics, trends, and velocity calculations
    """
    try:
        if not os.path.exists(RESULTS_DIR):
            return {
                "runs": [],
                "trends": {},
                "velocity": {},
                "summary": {}
            }
        
        # Collect all result files
        result_files = []
        for file in os.listdir(RESULTS_DIR):
            if not file.endswith(".json"):
                continue
            
            file_path = os.path.join(RESULTS_DIR, file)
            try:
                with open(file_path, 'r') as f:
                    data = json.load(f)
                    
                    # Filter by pipeline name if specified
                    if pipeline_name and data.get('pipeline_name') != pipeline_name:
                        continue
                    
                    result_files.append({
                        "file_path": file_path,
                        "run_id": data.get('run_id'),
                        "pipeline_name": data.get('pipeline_name'),
                        "started_at": data.get('started_at'),
                        "data": data
                    })
            except Exception as e:
                continue
        
        # Sort by timestamp (newest first)
        result_files.sort(key=lambda x: x['started_at'] or '', reverse=True)
        
        # Limit number of runs
        result_files = result_files[:limit]
        
        # Reverse for chronological order (oldest to newest) for trend calculation
        result_files.reverse()
        
        # Extract historical data points
        historical_runs = []
        for rf in result_files:
            data = rf['data']
            steps = data.get('steps', [])
            
            # Calculate metrics for this run
            total_steps = len(steps)
            passed_steps = sum(1 for s in steps if s.get('status') == 'success')
            failed_steps = sum(1 for s in steps if s.get('status') == 'failure')
            total_errors = sum(s.get('error_count', 0) for s in steps)
            
            # Count severity levels
            blocker_count = sum(1 for s in steps if s.get('severity') == 'BLOCKER')
            high_count = sum(1 for s in steps if s.get('severity') == 'HIGH')
            medium_count = sum(1 for s in steps if s.get('severity') == 'MEDIUM')
            
            # Calculate success rate
            success_rate = (passed_steps / total_steps * 100) if total_steps > 0 else 0
            
            historical_runs.append({
                "run_id": rf['run_id'],
                "pipeline_name": rf['pipeline_name'],
                "timestamp": rf['started_at'],
                "metrics": {
                    "total_steps": total_steps,
                    "passed_steps": passed_steps,
                    "failed_steps": failed_steps,
                    "success_rate": round(success_rate, 2),
                    "total_errors": total_errors,
                    "blocker_issues": blocker_count,
                    "high_severity_issues": high_count,
                    "medium_severity_issues": medium_count
                }
            })
        
        # Calculate trends
        trends = _calculate_trends(historical_runs)
        
        # Calculate velocity (rate of improvement)
        velocity = _calculate_velocity(historical_runs)
        
        # Summary statistics
        summary = _calculate_historical_summary(historical_runs)
        
        return {
            "runs": historical_runs,
            "trends": trends,
            "velocity": velocity,
            "summary": summary
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching historical trends: {str(e)}")


def _calculate_trends(runs: List[Dict]) -> Dict:
    """
    Calculate trend indicators for key metrics.
    
    Returns:
        Trend direction and change percentages for each metric
    """
    if len(runs) < 2:
        return {
            "success_rate": {"trend": "stable", "change_percent": 0},
            "total_errors": {"trend": "stable", "change_percent": 0},
            "blocker_issues": {"trend": "stable", "change_percent": 0}
        }
    
    # Compare most recent run to oldest run
    oldest = runs[0]['metrics']
    newest = runs[-1]['metrics']
    
    # Success rate trend
    success_rate_change = newest['success_rate'] - oldest['success_rate']
    success_rate_trend = "improving" if success_rate_change > 5 else ("degrading" if success_rate_change < -5 else "stable")
    
    # Total errors trend (lower is better)
    if oldest['total_errors'] > 0:
        error_change_percent = ((newest['total_errors'] - oldest['total_errors']) / oldest['total_errors']) * 100
    else:
        error_change_percent = 0 if newest['total_errors'] == 0 else 100
    
    error_trend = "improving" if error_change_percent < -10 else ("degrading" if error_change_percent > 10 else "stable")
    
    # Blocker issues trend (lower is better)
    if oldest['blocker_issues'] > 0:
        blocker_change_percent = ((newest['blocker_issues'] - oldest['blocker_issues']) / oldest['blocker_issues']) * 100
    else:
        blocker_change_percent = 0 if newest['blocker_issues'] == 0 else 100
    
    blocker_trend = "improving" if blocker_change_percent < -10 else ("degrading" if blocker_change_percent > 10 else "stable")
    
    return {
        "success_rate": {
            "trend": success_rate_trend,
            "change_percent": round(success_rate_change, 2),
            "direction": "up" if success_rate_change > 0 else ("down" if success_rate_change < 0 else "neutral")
        },
        "total_errors": {
            "trend": error_trend,
            "change_percent": round(error_change_percent, 2),
            "direction": "down" if error_change_percent < 0 else ("up" if error_change_percent > 0 else "neutral")
        },
        "blocker_issues": {
            "trend": blocker_trend,
            "change_percent": round(blocker_change_percent, 2),
            "direction": "down" if blocker_change_percent < 0 else ("up" if blocker_change_percent > 0 else "neutral")
        }
    }


def _calculate_velocity(runs: List[Dict]) -> Dict:
    """
    Calculate velocity metrics showing rate of improvement.
    
    Returns:
        Velocity indicators and projected timelines
    """
    if len(runs) < 3:
        return {
            "average_improvement_rate": 0,
            "velocity_indicator": "insufficient_data",
            "estimated_runs_to_100_percent": None
        }
    
    # Calculate moving average of success rates
    success_rates = [r['metrics']['success_rate'] for r in runs]
    
    # Calculate average improvement per run
    improvements = []
    for i in range(1, len(success_rates)):
        improvements.append(success_rates[i] - success_rates[i-1])
    
    avg_improvement = sum(improvements) / len(improvements) if improvements else 0
    
    # Determine velocity indicator
    if avg_improvement > 5:
        velocity_indicator = "accelerating"
    elif avg_improvement > 2:
        velocity_indicator = "steady"
    elif avg_improvement > -2:
        velocity_indicator = "slow"
    else:
        velocity_indicator = "degrading"
    
    # Estimate runs to 100% success
    current_success_rate = runs[-1]['metrics']['success_rate']
    if avg_improvement > 0 and current_success_rate < 100:
        estimated_runs = int((100 - current_success_rate) / avg_improvement)
    else:
        estimated_runs = None
    
    return {
        "average_improvement_rate": round(avg_improvement, 2),
        "velocity_indicator": velocity_indicator,
        "estimated_runs_to_100_percent": estimated_runs,
        "current_success_rate": round(current_success_rate, 2)
    }


def _calculate_historical_summary(runs: List[Dict]) -> Dict:
    """
    Calculate summary statistics across all runs.
    
    Returns:
        Summary metrics and milestones
    """
    if not runs:
        return {
            "total_runs": 0,
            "best_run": None,
            "worst_run": None,
            "average_success_rate": 0,
            "total_issues_resolved": 0
        }
    
    # Find best and worst runs
    best_run = max(runs, key=lambda r: r['metrics']['success_rate'])
    worst_run = min(runs, key=lambda r: r['metrics']['success_rate'])
    
    # Calculate averages
    avg_success_rate = sum(r['metrics']['success_rate'] for r in runs) / len(runs)
    
    # Calculate total issues resolved (if trending positive)
    if len(runs) >= 2:
        initial_errors = runs[0]['metrics']['total_errors']
        current_errors = runs[-1]['metrics']['total_errors']
        issues_resolved = max(0, initial_errors - current_errors)
    else:
        issues_resolved = 0
    
    return {
        "total_runs": len(runs),
        "best_run": {
            "run_id": best_run['run_id'],
            "success_rate": best_run['metrics']['success_rate'],
            "timestamp": best_run['timestamp']
        },
        "worst_run": {
            "run_id": worst_run['run_id'],
            "success_rate": worst_run['metrics']['success_rate'],
            "timestamp": worst_run['timestamp']
        },
        "average_success_rate": round(avg_success_rate, 2),
        "total_issues_resolved": issues_resolved
    }


# ============================================================================
# BASELINE MANAGEMENT ENDPOINTS
# ============================================================================

@router.post("/baseline/set")
def set_baseline(request: Dict[str, str]):
    """
    Set a specific run as the baseline for comparison.

    Request body:
        {
            "run_id": "run_20250104_123456",
            "pipeline_name": "my_pipeline"  # optional
        }
    """
    try:
        run_id = request.get("run_id")
        if not run_id:
            raise HTTPException(status_code=400, detail="run_id is required")

        # Find the result file
        result_file = os.path.join(RESULTS_DIR, f"{run_id}.json")
        if not os.path.exists(result_file):
            raise HTTPException(status_code=404, detail=f"Run {run_id} not found")

        # Load the result data
        with open(result_file, 'r') as f:
            result_data = json.load(f)

        # Extract key metrics
        pipeline_name = result_data.get('pipeline_name', request.get('pipeline_name', 'unknown'))
        timestamp = result_data.get('started_at') or result_data.get('timestamp') or result_data.get('execution_time', '')

        # Calculate metrics
        steps = result_data.get('steps', [])
        total_steps = len(steps)
        passed_steps = sum(1 for s in steps if s.get('status') == 'success')
        failed_steps = sum(1 for s in steps if s.get('status') == 'failure')
        success_rate = (passed_steps / total_steps * 100) if total_steps > 0 else 0

        # Count errors
        total_errors = 0
        blocker_issues = 0
        high_severity_issues = 0

        for step in steps:
            error_count = step.get('error_count', 0)
            total_errors += error_count

            severity = step.get('severity', '').upper()
            if severity == 'BLOCKER':
                blocker_issues += 1
            elif severity == 'HIGH':
                high_severity_issues += 1

        # Store baseline information
        baseline_data = {
            "run_id": run_id,
            "pipeline_name": pipeline_name,
            "timestamp": timestamp,
            "set_at": datetime.now().isoformat(),
            "metrics": {
                "total_steps": total_steps,
                "passed_steps": passed_steps,
                "failed_steps": failed_steps,
                "success_rate": round(success_rate, 2),
                "total_errors": total_errors,
                "blocker_issues": blocker_issues,
                "high_severity_issues": high_severity_issues
            }
        }

        # Save baseline to file
        with open(BASELINE_FILE, 'w') as f:
            json.dump(baseline_data, f, indent=2)

        return {
            "message": "Baseline set successfully",
            "baseline": baseline_data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to set baseline: {str(e)}")


@router.get("/baseline")
def get_baseline():
    """
    Get the current baseline run information.

    Returns the baseline data if set, otherwise returns null.
    """
    try:
        if not os.path.exists(BASELINE_FILE):
            return {"baseline": None}

        with open(BASELINE_FILE, 'r') as f:
            baseline_data = json.load(f)

        return {"baseline": baseline_data}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get baseline: {str(e)}")


@router.delete("/baseline")
def clear_baseline():
    """
    Clear the current baseline.
    """
    try:
        if os.path.exists(BASELINE_FILE):
            os.remove(BASELINE_FILE)

        return {"message": "Baseline cleared successfully"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to clear baseline: {str(e)}")


@router.get("/baseline/compare/{run_id}")
def compare_to_baseline(run_id: str):
    """
    Compare a specific run against the current baseline.

    Returns comparison metrics showing improvements/regressions.
    """
    try:
        # Check if baseline exists
        if not os.path.exists(BASELINE_FILE):
            raise HTTPException(status_code=404, detail="No baseline set. Please set a baseline first.")

        # Load baseline
        with open(BASELINE_FILE, 'r') as f:
            baseline_data = json.load(f)

        # Find the comparison run
        result_file = os.path.join(RESULTS_DIR, f"{run_id}.json")
        if not os.path.exists(result_file):
            raise HTTPException(status_code=404, detail=f"Run {run_id} not found")

        # Load the comparison run data
        with open(result_file, 'r') as f:
            result_data = json.load(f)

        # Extract metrics from comparison run
        steps = result_data.get('steps', [])
        total_steps = len(steps)
        passed_steps = sum(1 for s in steps if s.get('status') == 'success')
        failed_steps = sum(1 for s in steps if s.get('status') == 'failure')
        success_rate = (passed_steps / total_steps * 100) if total_steps > 0 else 0

        total_errors = 0
        blocker_issues = 0
        high_severity_issues = 0

        for step in steps:
            error_count = step.get('error_count', 0)
            total_errors += error_count

            severity = step.get('severity', '').upper()
            if severity == 'BLOCKER':
                blocker_issues += 1
            elif severity == 'HIGH':
                high_severity_issues += 1

        comparison_metrics = {
            "total_steps": total_steps,
            "passed_steps": passed_steps,
            "failed_steps": failed_steps,
            "success_rate": round(success_rate, 2),
            "total_errors": total_errors,
            "blocker_issues": blocker_issues,
            "high_severity_issues": high_severity_issues
        }

        # Calculate deltas
        baseline_metrics = baseline_data['metrics']

        def calculate_delta(current, baseline_val, key):
            delta = current - baseline_val
            if key in ['success_rate']:
                # For success rate, positive is good
                status = 'improved' if delta > 0 else 'degraded' if delta < 0 else 'unchanged'
            else:
                # For errors/issues, negative is good (fewer issues)
                status = 'improved' if delta < 0 else 'degraded' if delta > 0 else 'unchanged'

            return {
                "baseline": baseline_val,
                "current": current,
                "delta": round(delta, 2),
                "delta_percent": round((delta / baseline_val * 100) if baseline_val != 0 else 0, 2),
                "status": status
            }

        deltas = {
            "success_rate": calculate_delta(
                comparison_metrics['success_rate'],
                baseline_metrics['success_rate'],
                'success_rate'
            ),
            "total_errors": calculate_delta(
                comparison_metrics['total_errors'],
                baseline_metrics['total_errors'],
                'total_errors'
            ),
            "blocker_issues": calculate_delta(
                comparison_metrics['blocker_issues'],
                baseline_metrics['blocker_issues'],
                'blocker_issues'
            ),
            "high_severity_issues": calculate_delta(
                comparison_metrics['high_severity_issues'],
                baseline_metrics['high_severity_issues'],
                'high_severity_issues'
            ),
            "passed_steps": calculate_delta(
                comparison_metrics['passed_steps'],
                baseline_metrics['passed_steps'],
                'passed_steps_inverted'  # More passed steps is good
            ),
            "failed_steps": calculate_delta(
                comparison_metrics['failed_steps'],
                baseline_metrics['failed_steps'],
                'failed_steps'
            )
        }

        # Overall assessment
        improved_count = sum(1 for d in deltas.values() if d['status'] == 'improved')
        degraded_count = sum(1 for d in deltas.values() if d['status'] == 'degraded')

        if improved_count > degraded_count:
            overall_status = 'improved'
        elif degraded_count > improved_count:
            overall_status = 'degraded'
        else:
            overall_status = 'mixed'

        return {
            "baseline": baseline_data,
            "comparison_run": {
                "run_id": run_id,
                "pipeline_name": result_data.get('pipeline_name', 'unknown'),
                "timestamp": result_data.get('started_at') or result_data.get('timestamp') or result_data.get('execution_time', ''),
                "metrics": comparison_metrics
            },
            "deltas": deltas,
            "overall_status": overall_status,
            "summary": {
                "improved_metrics": improved_count,
                "degraded_metrics": degraded_count,
                "unchanged_metrics": len(deltas) - improved_count - degraded_count
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to compare to baseline: {str(e)}")

# ============================================================================
# EXPORT ENDPOINTS
# ============================================================================

@router.get("/export/json/{run_id}")
def export_json(run_id: str):
    """Export validation results as JSON format."""
    try:
        result_file = os.path.join(RESULTS_DIR, f"{run_id}.json")
        
        if not os.path.exists(result_file):
            raise HTTPException(status_code=404, detail=f"Results file not found for run_id: {run_id}")
        
        with open(result_file, 'r') as f:
            results = json.load(f)
        
        # Create a formatted JSON export
        export_data = {
            "run_id": run_id,
            "exported_at": datetime.now().isoformat(),
            "pipeline_name": results.get("pipeline_name", "Unknown"),
            "timestamp": results.get("timestamp"),
            "summary": {
                "total_steps": results.get("total_steps", 0),
                "passed_steps": results.get("passed_steps", 0),
                "failed_steps": results.get("failed_steps", 0),
                "success_rate": results.get("success_rate", 0),
                "total_errors": results.get("total_errors", 0)
            },
            "steps": results.get("steps", [])
        }
        
        # Create JSON string
        json_str = json.dumps(export_data, indent=2)
        
        # Return as downloadable file
        return StreamingResponse(
            iter([json_str]),
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=validation_results_{run_id}.json"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export JSON: {str(e)}")


@router.get("/export/excel/{run_id}")
def export_excel(run_id: str):
    """Export validation results as Excel format."""
    try:
        result_file = os.path.join(RESULTS_DIR, f"{run_id}.json")
        
        if not os.path.exists(result_file):
            raise HTTPException(status_code=404, detail=f"Results file not found for run_id: {run_id}")
        
        with open(result_file, 'r') as f:
            results = json.load(f)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header styling
        header_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws_summary['A1'] = "Validation Results Summary"
        ws_summary['A1'].font = Font(name='Arial', size=16, bold=True)
        ws_summary.merge_cells('A1:B1')
        
        # Run information
        row = 3
        info_data = [
            ("Run ID:", run_id),
            ("Pipeline Name:", results.get("pipeline_name", "Unknown")),
            ("Timestamp:", results.get("timestamp", "Unknown")),
            ("Total Steps:", results.get("total_steps", 0)),
            ("Passed Steps:", results.get("passed_steps", 0)),
            ("Failed Steps:", results.get("failed_steps", 0)),
            ("Success Rate:", f"{results.get('success_rate', 0):.1f}%"),
            ("Total Errors:", results.get("total_errors", 0))
        ]
        
        for label, value in info_data:
            ws_summary[f'A{row}'] = label
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = value
            row += 1
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 40
        
        # Steps Sheet
        ws_steps = wb.create_sheet("Steps")
        
        # Headers
        headers = ["Step Name", "Status", "Severity", "Validation Type", "Message", "Error Count", "Execution Time"]
        for col, header in enumerate(headers, 1):
            cell = ws_steps.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        steps = results.get("steps", [])
        for row_idx, step in enumerate(steps, 2):
            ws_steps.cell(row=row_idx, column=1, value=step.get("step_name", ""))
            
            status = step.get("status", "unknown")
            status_cell = ws_steps.cell(row=row_idx, column=2, value=status)
            # Color code status
            if status == "success":
                status_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif status == "failure":
                status_cell.fill = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")
            elif status == "warning":
                status_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            
            ws_steps.cell(row=row_idx, column=3, value=step.get("severity", "N/A"))
            ws_steps.cell(row=row_idx, column=4, value=step.get("validation_type", ""))
            ws_steps.cell(row=row_idx, column=5, value=step.get("message", ""))
            ws_steps.cell(row=row_idx, column=6, value=step.get("error_count", 0))
            ws_steps.cell(row=row_idx, column=7, value=step.get("execution_time", ""))
        
        # Adjust column widths
        ws_steps.column_dimensions['A'].width = 30
        ws_steps.column_dimensions['B'].width = 12
        ws_steps.column_dimensions['C'].width = 12
        ws_steps.column_dimensions['D'].width = 20
        ws_steps.column_dimensions['E'].width = 50
        ws_steps.column_dimensions['F'].width = 12
        ws_steps.column_dimensions['G'].width = 15
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Return as downloadable file
        return StreamingResponse(
            iter([excel_file.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=validation_results_{run_id}.xlsx"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export Excel: {str(e)}")


@router.get("/export/pdf/{run_id}")
def export_pdf(run_id: str):
    """Export validation results as PDF format."""
    try:
        result_file = os.path.join(RESULTS_DIR, f"{run_id}.json")
        
        if not os.path.exists(result_file):
            raise HTTPException(status_code=404, detail=f"Results file not found for run_id: {run_id}")
        
        with open(result_file, 'r') as f:
            results = json.load(f)
        
        # Create PDF buffer
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter,
                              rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        elements.append(Paragraph("Validation Results Report", title_style))
        elements.append(Spacer(1, 12))
        
        # Summary Information
        elements.append(Paragraph("Summary", heading_style))
        
        summary_data = [
            ["Run ID:", run_id],
            ["Pipeline Name:", results.get("pipeline_name", "Unknown")],
            ["Timestamp:", results.get("timestamp", "Unknown")],
            ["Total Steps:", str(results.get("total_steps", 0))],
            ["Passed Steps:", str(results.get("passed_steps", 0))],
            ["Failed Steps:", str(results.get("failed_steps", 0))],
            ["Success Rate:", f"{results.get('success_rate', 0):.1f}%"],
            ["Total Errors:", str(results.get("total_errors", 0))]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E6E6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Steps Details
        elements.append(Paragraph("Validation Steps", heading_style))
        
        steps = results.get("steps", [])
        
        # Group steps by status
        passed_steps = [s for s in steps if s.get("status") == "success"]
        failed_steps = [s for s in steps if s.get("status") == "failure"]
        warning_steps = [s for s in steps if s.get("status") == "warning"]
        
        # Status Summary
        status_summary_data = [
            ["Status", "Count"],
            ["Passed", str(len(passed_steps))],
            ["Failed", str(len(failed_steps))],
            ["Warning", str(len(warning_steps))]
        ]
        
        status_table = Table(status_summary_data, colWidths=[2*inch, 2*inch])
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(status_table)
        elements.append(Spacer(1, 20))
        
        # Detailed Steps Table
        steps_header = ["Step Name", "Status", "Severity", "Errors"]
        steps_data = [steps_header]
        
        for step in steps:
            status = step.get("status", "unknown")
            steps_data.append([
                Paragraph(step.get("step_name", "")[:40], styles['Normal']),
                status.upper(),
                step.get("severity", "N/A"),
                str(step.get("error_count", 0))
            ])
        
        steps_table = Table(steps_data, colWidths=[3*inch, 1*inch, 1*inch, 0.8*inch])
        steps_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        # Color code by status
        for i, step in enumerate(steps, 1):
            status = step.get("status", "unknown")
            if status == "success":
                steps_table.setStyle(TableStyle([('BACKGROUND', (1, i), (1, i), colors.HexColor('#C6E0B4'))]))
            elif status == "failure":
                steps_table.setStyle(TableStyle([('BACKGROUND', (1, i), (1, i), colors.HexColor('#F4C7C3'))]))
            elif status == "warning":
                steps_table.setStyle(TableStyle([('BACKGROUND', (1, i), (1, i), colors.HexColor('#FFE699'))]))
        
        elements.append(steps_table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF bytes
        pdf_buffer.seek(0)
        
        # Return as downloadable file
        return StreamingResponse(
            iter([pdf_buffer.getvalue()]),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=validation_results_{run_id}.pdf"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export PDF: {str(e)}")


# ============================================================================
# BATCH EXPORT ENDPOINTS
# ============================================================================

@router.get("/export/json/batch/{batch_id}")
def export_batch_json(batch_id: str):
    """Export batch validation results as JSON format."""
    try:
        result_file = os.path.join(RESULTS_DIR, f"{batch_id}.json")

        if not os.path.exists(result_file):
            raise HTTPException(status_code=404, detail=f"Batch results file not found for batch_id: {batch_id}")

        with open(result_file, 'r') as f:
            batch_results = json.load(f)

        # Create a formatted JSON export
        export_data = {
            "batch_id": batch_id,
            "exported_at": datetime.now().isoformat(),
            "batch_job_name": batch_results.get("batch_job_name", "Unknown"),
            "pipeline_name": batch_results.get("pipeline_name", "Unknown"),
            "timestamp": batch_results.get("started_at"),
            "status": batch_results.get("status", "UNKNOWN"),
            "summary": {
                "total_pipelines": batch_results.get("total_pipelines", 0),
                "total_validations": batch_results.get("summary", {}).get("total_validations", 0),
                "passed": batch_results.get("summary", {}).get("passed", 0),
                "failed": batch_results.get("summary", {}).get("failed", 0),
                "pass_rate": batch_results.get("summary", {}).get("pass_rate", 0),
                "tables_validated": batch_results.get("tables_validated", [])
            },
            "tables": batch_results.get("tables", []),
            "results": batch_results.get("results", []),
            "individual_run_ids": batch_results.get("individual_run_ids", [])
        }

        # Create JSON string
        json_str = json.dumps(export_data, indent=2)

        # Return as downloadable file
        return StreamingResponse(
            iter([json_str]),
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=batch_validation_results_{batch_id}.json"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export batch JSON: {str(e)}")


@router.get("/export/excel/batch/{batch_id}")
def export_batch_excel(batch_id: str):
    """Export batch validation results as Excel format."""
    try:
        result_file = os.path.join(RESULTS_DIR, f"{batch_id}.json")

        if not os.path.exists(result_file):
            raise HTTPException(status_code=404, detail=f"Batch results file not found for batch_id: {batch_id}")

        with open(result_file, 'r') as f:
            batch_results = json.load(f)

        # Create Excel workbook
        wb = openpyxl.Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Batch Summary"

        # Header styling
        header_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Title
        ws_summary['A1'] = "Batch Validation Results Summary"
        ws_summary['A1'].font = Font(name='Arial', size=16, bold=True)
        ws_summary.merge_cells('A1:B1')

        # Batch information
        row = 3
        summary = batch_results.get("summary", {})
        info_data = [
            ("Batch ID:", batch_id),
            ("Batch Job Name:", batch_results.get("batch_job_name", "Unknown")),
            ("Pipeline Name:", batch_results.get("pipeline_name", "Unknown")),
            ("Status:", batch_results.get("status", "UNKNOWN")),
            ("Started At:", batch_results.get("started_at", "Unknown")),
            ("Total Pipelines:", batch_results.get("total_pipelines", 0)),
            ("Total Validations:", summary.get("total_validations", 0)),
            ("Passed:", summary.get("passed", 0)),
            ("Failed:", summary.get("failed", 0)),
            ("Pass Rate:", f"{summary.get('pass_rate', 0):.1f}%")
        ]

        for label, value in info_data:
            ws_summary[f'A{row}'] = label
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = value
            row += 1

        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 50

        # Tables Sheet
        ws_tables = wb.create_sheet("Tables Validated")

        # Headers
        table_headers = ["Table", "Schema", "Table Name", "Total Validations", "Passed", "Failed"]
        for col, header in enumerate(table_headers, 1):
            cell = ws_tables.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        tables = batch_results.get("tables", [])
        for row_idx, table in enumerate(tables, 2):
            ws_tables.cell(row=row_idx, column=1, value=table.get("table", ""))
            ws_tables.cell(row=row_idx, column=2, value=table.get("schema", ""))
            ws_tables.cell(row=row_idx, column=3, value=table.get("table_name", ""))
            ws_tables.cell(row=row_idx, column=4, value=table.get("total_validations", 0))
            ws_tables.cell(row=row_idx, column=5, value=table.get("passed", 0))
            ws_tables.cell(row=row_idx, column=6, value=table.get("failed", 0))

        # Adjust column widths
        ws_tables.column_dimensions['A'].width = 30
        ws_tables.column_dimensions['B'].width = 15
        ws_tables.column_dimensions['C'].width = 20
        ws_tables.column_dimensions['D'].width = 18
        ws_tables.column_dimensions['E'].width = 12
        ws_tables.column_dimensions['F'].width = 12

        # Validation Results Sheet
        ws_results = wb.create_sheet("Validation Results")

        # Headers
        result_headers = ["Validation Name", "Status", "Severity", "Error Message", "Timestamp"]
        for col, header in enumerate(result_headers, 1):
            cell = ws_results.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        results = batch_results.get("results", [])
        for row_idx, result in enumerate(results, 2):
            ws_results.cell(row=row_idx, column=1, value=result.get("name", ""))

            status = result.get("status", "unknown")
            status_cell = ws_results.cell(row=row_idx, column=2, value=status)
            # Color code status
            if status == "PASS":
                status_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif status == "FAIL":
                status_cell.fill = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")
            elif status == "ERROR":
                status_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

            ws_results.cell(row=row_idx, column=3, value=result.get("severity", "N/A"))

            # Handle error details
            details = result.get("details", {})
            error_msg = details.get("error", "") if isinstance(details, dict) else str(details)
            ws_results.cell(row=row_idx, column=4, value=error_msg[:200])  # Truncate long errors

            ws_results.cell(row=row_idx, column=5, value=result.get("timestamp", ""))

        # Adjust column widths
        ws_results.column_dimensions['A'].width = 35
        ws_results.column_dimensions['B'].width = 12
        ws_results.column_dimensions['C'].width = 12
        ws_results.column_dimensions['D'].width = 60
        ws_results.column_dimensions['E'].width = 25

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return as downloadable file
        return StreamingResponse(
            iter([excel_file.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=batch_validation_results_{batch_id}.xlsx"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export batch Excel: {str(e)}")


@router.get("/export/pdf/batch/{batch_id}")
def export_batch_pdf(batch_id: str):
    """Export batch validation results as PDF format."""
    try:
        result_file = os.path.join(RESULTS_DIR, f"{batch_id}.json")

        if not os.path.exists(result_file):
            raise HTTPException(status_code=404, detail=f"Batch results file not found for batch_id: {batch_id}")

        with open(result_file, 'r') as f:
            batch_results = json.load(f)

        # Create PDF buffer
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter,
                              rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)

        # Container for PDF elements
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=12,
            spaceBefore=12
        )

        # Title
        elements.append(Paragraph("Batch Validation Results Report", title_style))
        elements.append(Spacer(1, 12))

        # Summary Information
        elements.append(Paragraph("Batch Summary", heading_style))

        summary = batch_results.get("summary", {})
        summary_data = [
            ["Batch ID:", batch_id],
            ["Batch Job Name:", batch_results.get("batch_job_name", "Unknown")],
            ["Pipeline Name:", batch_results.get("pipeline_name", "Unknown")],
            ["Status:", batch_results.get("status", "UNKNOWN")],
            ["Started At:", batch_results.get("started_at", "Unknown")],
            ["Total Pipelines:", str(batch_results.get("total_pipelines", 0))],
            ["Total Validations:", str(summary.get("total_validations", 0))],
            ["Passed:", str(summary.get("passed", 0))],
            ["Failed:", str(summary.get("failed", 0))],
            ["Pass Rate:", f"{summary.get('pass_rate', 0):.1f}%"]
        ]

        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E6E6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Tables Validated
        elements.append(Paragraph("Tables Validated", heading_style))

        tables = batch_results.get("tables", [])
        if tables:
            table_headers = ["Table", "Total", "Passed", "Failed"]
            table_data = [table_headers]

            for table in tables:
                table_data.append([
                    Paragraph(table.get("table", "")[:30], styles['Normal']),
                    str(table.get("total_validations", 0)),
                    str(table.get("passed", 0)),
                    str(table.get("failed", 0))
                ])

            tables_table = Table(table_data, colWidths=[3*inch, 1*inch, 1*inch, 1*inch])
            tables_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))

            elements.append(tables_table)
        else:
            elements.append(Paragraph("No tables validated", styles['Normal']))

        elements.append(Spacer(1, 20))

        # Validation Results Summary
        elements.append(Paragraph("Validation Results", heading_style))

        results = batch_results.get("results", [])

        # Count by status
        passed_count = len([r for r in results if r.get("status") == "PASS"])
        failed_count = len([r for r in results if r.get("status") == "FAIL"])
        error_count = len([r for r in results if r.get("status") == "ERROR"])

        status_summary_data = [
            ["Status", "Count"],
            ["Passed", str(passed_count)],
            ["Failed", str(failed_count)],
            ["Error", str(error_count)]
        ]

        status_table = Table(status_summary_data, colWidths=[2*inch, 2*inch])
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(status_table)
        elements.append(Spacer(1, 20))

        # Detailed Results Table (only show errors and failures)
        if results:
            elements.append(Paragraph("Validation Details (Errors & Failures)", heading_style))

            results_header = ["Validation Name", "Status", "Severity", "Error"]
            results_data = [results_header]

            # Only show failed and error results
            failed_results = [r for r in results if r.get("status") in ["FAIL", "ERROR"]]

            if failed_results:
                for result in failed_results[:20]:  # Limit to 20 to prevent huge PDFs
                    status = result.get("status", "unknown")
                    details = result.get("details", {})
                    error_msg = details.get("error", "") if isinstance(details, dict) else str(details)

                    results_data.append([
                        Paragraph(result.get("name", "")[:30], styles['Normal']),
                        status,
                        result.get("severity", "N/A"),
                        Paragraph(error_msg[:100], styles['Normal'])  # Truncate long errors
                    ])

                results_table = Table(results_data, colWidths=[2*inch, 0.8*inch, 0.8*inch, 2.4*inch])
                results_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))

                # Color code by status
                for i, result in enumerate(failed_results[:20], 1):
                    status = result.get("status", "unknown")
                    if status == "FAIL":
                        results_table.setStyle(TableStyle([('BACKGROUND', (1, i), (1, i), colors.HexColor('#F4C7C3'))]))
                    elif status == "ERROR":
                        results_table.setStyle(TableStyle([('BACKGROUND', (1, i), (1, i), colors.HexColor('#FFE699'))]))

                elements.append(results_table)
            else:
                elements.append(Paragraph("All validations passed!", styles['Normal']))

        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))

        # Build PDF
        doc.build(elements)

        # Get PDF bytes
        pdf_buffer.seek(0)

        # Return as downloadable file
        return StreamingResponse(
            iter([pdf_buffer.getvalue()]),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=batch_validation_results_{batch_id}.pdf"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export batch PDF: {str(e)}")


# ============================================================================
# COMPARISON EXPORT ENDPOINTS (with Analysis Data)
# ============================================================================

@router.post("/export/json/comparison")
async def export_comparison_json(comparison_data: dict):
    """
    Export comparison data (with executive summary, trends, etc.) as JSON
    """
    try:
        # Add metadata
        export_data = {
            "exported_at": datetime.now().isoformat(),
            "export_type": "migration_readiness_comparison",
            **comparison_data
        }

        # Convert to JSON string
        json_str = json.dumps(export_data, indent=2, default=str)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"migration_readiness_report_{timestamp}.json"

        # Return as downloadable file
        return StreamingResponse(
            iter([json_str.encode()]),
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export comparison JSON: {str(e)}")


@router.post("/export/excel/comparison")
async def export_comparison_excel(comparison_data: dict):
    """
    Export comparison data (with executive summary, trends, etc.) as multi-sheet Excel workbook
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Helper function to apply header style
        def style_header_row(ws, row_num, max_col):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # Helper function to auto-size columns
        def auto_size_columns(ws):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # ==================== SHEET 1: Executive Summary ====================
        if 'executive_summary' in comparison_data:
            ws_summary = wb.create_sheet("Executive Summary")
            exec_sum = comparison_data['executive_summary']

            # Title
            ws_summary['A1'] = "Migration Readiness Report - Executive Summary"
            ws_summary['A1'].font = Font(bold=True, size=14)
            ws_summary.merge_cells('A1:B1')

            # Metadata
            row = 3
            ws_summary[f'A{row}'] = "Generated:"
            ws_summary[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row += 2

            # Readiness Score
            ws_summary[f'A{row}'] = "READINESS SCORE"
            ws_summary[f'A{row}'].font = section_font
            ws_summary[f'A{row}'].fill = section_fill
            row += 1

            ws_summary[f'A{row}'] = "Score:"
            ws_summary[f'B{row}'] = exec_sum.get('readiness_score', 0)
            row += 1
            ws_summary[f'A{row}'] = "Overall Status:"
            ws_summary[f'B{row}'] = exec_sum.get('overall_status', 'N/A')
            row += 2

            # Validation Summary
            ws_summary[f'A{row}'] = "VALIDATION SUMMARY"
            ws_summary[f'A{row}'].font = section_font
            ws_summary[f'A{row}'].fill = section_fill
            row += 1

            ws_summary[f'A{row}'] = "Total Validations:"
            ws_summary[f'B{row}'] = exec_sum.get('total_validations', 0)
            row += 1
            ws_summary[f'A{row}'] = "Passed:"
            ws_summary[f'B{row}'] = exec_sum.get('passed_validations', 0)
            row += 1
            ws_summary[f'A{row}'] = "Failed:"
            ws_summary[f'B{row}'] = exec_sum.get('failed_validations', 0)
            row += 1
            ws_summary[f'A{row}'] = "Warnings:"
            ws_summary[f'B{row}'] = exec_sum.get('warnings', 0)
            row += 1
            ws_summary[f'A{row}'] = "Critical Issues:"
            ws_summary[f'B{row}'] = exec_sum.get('critical_issues', 0)
            row += 2

            # Severity Breakdown
            if 'severity_breakdown' in exec_sum:
                ws_summary[f'A{row}'] = "SEVERITY BREAKDOWN"
                ws_summary[f'A{row}'].font = section_font
                ws_summary[f'A{row}'].fill = section_fill
                row += 1

                sev = exec_sum['severity_breakdown']
                ws_summary[f'A{row}'] = "Blocker:"
                ws_summary[f'B{row}'] = sev.get('BLOCKER', 0)
                row += 1
                ws_summary[f'A{row}'] = "High:"
                ws_summary[f'B{row}'] = sev.get('HIGH', 0)
                row += 1
                ws_summary[f'A{row}'] = "Medium:"
                ws_summary[f'B{row}'] = sev.get('MEDIUM', 0)
                row += 1
                ws_summary[f'A{row}'] = "Low:"
                ws_summary[f'B{row}'] = sev.get('LOW', 0)

            auto_size_columns(ws_summary)

        # ==================== SHEET 2: Trend Analysis ====================
        if 'trend_analysis' in comparison_data:
            ws_trends = wb.create_sheet("Trend Analysis")
            trends = comparison_data['trend_analysis']

            # Title
            ws_trends['A1'] = "Trend Analysis"
            ws_trends['A1'].font = Font(bold=True, size=14)
            ws_trends.merge_cells('A1:C1')

            row = 3

            # Velocity
            if 'velocity' in trends:
                ws_trends[f'A{row}'] = "VELOCITY METRICS"
                ws_trends[f'A{row}'].font = section_font
                ws_trends[f'A{row}'].fill = section_fill
                row += 1

                vel = trends['velocity']
                ws_trends[f'A{row}'] = "Per Day:"
                ws_trends[f'B{row}'] = vel.get('per_day', 0)
                row += 1
                ws_trends[f'A{row}'] = "Per Week:"
                ws_trends[f'B{row}'] = vel.get('per_week', 0)
                row += 2

            # Projections
            ws_trends[f'A{row}'] = "PROJECTIONS"
            ws_trends[f'A{row}'].font = section_font
            ws_trends[f'A{row}'].fill = section_fill
            row += 1

            ws_trends[f'A{row}'] = "Projected Zero Date:"
            ws_trends[f'B{row}'] = trends.get('projected_zero_date', 'N/A')
            row += 1
            ws_trends[f'A{row}'] = "Regression Detected:"
            ws_trends[f'B{row}'] = str(trends.get('regression_detected', False))
            row += 2

            # Error Trend Table
            if 'error_trend' in trends and trends['error_trend']:
                ws_trends[f'A{row}'] = "ERROR TREND HISTORY"
                ws_trends[f'A{row}'].font = section_font
                ws_trends[f'A{row}'].fill = section_fill
                row += 1

                # Headers
                ws_trends[f'A{row}'] = "Timestamp"
                ws_trends[f'B{row}'] = "Total Errors"
                style_header_row(ws_trends, row, 2)
                row += 1

                # Data
                for trend_point in trends['error_trend']:
                    ws_trends[f'A{row}'] = trend_point.get('timestamp', '')
                    ws_trends[f'B{row}'] = trend_point.get('total_errors', 0)
                    row += 1

            auto_size_columns(ws_trends)

        # ==================== SHEET 3: Root Causes ====================
        if 'root_cause_groups' in comparison_data and comparison_data['root_cause_groups']:
            ws_root = wb.create_sheet("Root Causes")

            # Title
            ws_root['A1'] = "Root Cause Analysis"
            ws_root['A1'].font = Font(bold=True, size=14)
            ws_root.merge_cells('A1:E1')

            row = 3

            # Headers
            ws_root[f'A{row}'] = "Category"
            ws_root[f'B{row}'] = "Title"
            ws_root[f'C{row}'] = "Severity"
            ws_root[f'D{row}'] = "Total Errors"
            ws_root[f'E{row}'] = "Affected Steps"
            ws_root[f'F{row}'] = "Description"
            ws_root[f'G{row}'] = "Recommended Action"
            style_header_row(ws_root, row, 7)
            row += 1

            # Data
            for group in comparison_data['root_cause_groups']:
                ws_root[f'A{row}'] = group.get('category', '')
                ws_root[f'B{row}'] = group.get('title', '')
                ws_root[f'C{row}'] = group.get('severity', '')
                ws_root[f'D{row}'] = group.get('total_errors', 0)

                # Join affected steps
                affected = group.get('affected_steps', [])
                ws_root[f'E{row}'] = ', '.join(affected[:5])  # Limit to first 5
                if len(affected) > 5:
                    ws_root[f'E{row}'] = ws_root[f'E{row}'].value + f" (+{len(affected)-5} more)"

                ws_root[f'F{row}'] = group.get('description', '')
                ws_root[f'G{row}'] = group.get('recommended_action', '')

                # Color code by severity
                severity = group.get('severity', '').upper()
                if severity == 'BLOCKER':
                    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    font = Font(color="FFFFFF")
                elif severity == 'HIGH':
                    fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    font = Font(color="000000")
                elif severity == 'MEDIUM':
                    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    font = Font(color="000000")
                else:
                    fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                    font = Font(color="000000")

                ws_root[f'C{row}'].fill = fill
                ws_root[f'C{row}'].font = font

                row += 1

            auto_size_columns(ws_root)

        # ==================== SHEET 4: Recommendations ====================
        if 'recommendations' in comparison_data and comparison_data['recommendations']:
            ws_rec = wb.create_sheet("Recommendations")

            # Title
            ws_rec['A1'] = "Prioritized Recommendations"
            ws_rec['A1'].font = Font(bold=True, size=14)
            ws_rec.merge_cells('A1:E1')

            row = 3

            # Headers
            ws_rec[f'A{row}'] = "Priority"
            ws_rec[f'B{row}'] = "Category"
            ws_rec[f'C{row}'] = "Title"
            ws_rec[f'D{row}'] = "Effort"
            ws_rec[f'E{row}'] = "Impact"
            ws_rec[f'F{row}'] = "Affected Count"
            ws_rec[f'G{row}'] = "Description"
            ws_rec[f'H{row}'] = "Action Items"
            style_header_row(ws_rec, row, 8)
            row += 1

            # Data
            for rec in comparison_data['recommendations']:
                ws_rec[f'A{row}'] = rec.get('priority', '')
                ws_rec[f'B{row}'] = rec.get('category', '')
                ws_rec[f'C{row}'] = rec.get('title', '')
                ws_rec[f'D{row}'] = rec.get('effort', '')
                ws_rec[f'E{row}'] = rec.get('impact', '')
                ws_rec[f'F{row}'] = rec.get('affected_count', 0)
                ws_rec[f'G{row}'] = rec.get('description', '')

                # Join action items
                action_items = rec.get('action_items', [])
                ws_rec[f'H{row}'] = '\n'.join([f"• {item}" for item in action_items])

                # Color code by priority
                priority = rec.get('priority', '').upper()
                if priority == 'CRITICAL':
                    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    font = Font(color="FFFFFF", bold=True)
                elif priority == 'HIGH':
                    fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    font = Font(color="000000", bold=True)
                elif priority == 'MEDIUM':
                    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    font = Font(color="000000")
                else:
                    fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                    font = Font(color="000000")

                ws_rec[f'A{row}'].fill = fill
                ws_rec[f'A{row}'].font = font

                # Enable text wrap for action items
                ws_rec[f'H{row}'].alignment = Alignment(wrap_text=True, vertical="top")

                row += 1

            auto_size_columns(ws_rec)

        # ==================== SHEET 5: Financial Impact ====================
        if 'financial_impact' in comparison_data:
            ws_fin = wb.create_sheet("Financial Impact")
            fin = comparison_data['financial_impact']

            # Title
            ws_fin['A1'] = "Financial Impact Assessment"
            ws_fin['A1'].font = Font(bold=True, size=14)
            ws_fin.merge_cells('A1:D1')

            row = 3

            # Table Criticality
            if 'table_criticality' in fin and fin['table_criticality']:
                ws_fin[f'A{row}'] = "TABLE CRITICALITY SCORES"
                ws_fin[f'A{row}'].font = section_font
                ws_fin[f'A{row}'].fill = section_fill
                row += 1

                # Headers
                ws_fin[f'A{row}'] = "Table"
                ws_fin[f'B{row}'] = "Criticality Score"
                ws_fin[f'C{row}'] = "Issues"
                ws_fin[f'D{row}'] = "Impact Level"
                style_header_row(ws_fin, row, 4)
                row += 1

                # Data
                for table in fin['table_criticality']:
                    ws_fin[f'A{row}'] = table.get('table', '')
                    ws_fin[f'B{row}'] = table.get('criticality_score', 0)
                    ws_fin[f'C{row}'] = table.get('issues', 0)
                    ws_fin[f'D{row}'] = table.get('impact_level', '')
                    row += 1

                row += 2

            # Financial Impact Summary
            if 'financial_impact' in fin and isinstance(fin['financial_impact'], dict):
                ws_fin[f'A{row}'] = "FINANCIAL IMPACT SUMMARY"
                ws_fin[f'A{row}'].font = section_font
                ws_fin[f'A{row}'].fill = section_fill
                row += 1

                impact_data = fin['financial_impact']
                for key, value in impact_data.items():
                    ws_fin[f'A{row}'] = key.replace('_', ' ').title() + ":"
                    ws_fin[f'B{row}'] = str(value)
                    row += 1

                row += 1

            # Risk Assessment
            if 'risk_assessment' in fin and isinstance(fin['risk_assessment'], dict):
                ws_fin[f'A{row}'] = "RISK ASSESSMENT"
                ws_fin[f'A{row}'].font = section_font
                ws_fin[f'A{row}'].fill = section_fill
                row += 1

                risk_data = fin['risk_assessment']
                for key, value in risk_data.items():
                    ws_fin[f'A{row}'] = key.replace('_', ' ').title() + ":"
                    ws_fin[f'B{row}'] = str(value)
                    row += 1

            auto_size_columns(ws_fin)

        # ==================== SHEET 6: Comparison Summary ====================
        if 'comparison_summary' in comparison_data:
            ws_comp_sum = wb.create_sheet("Comparison Summary")
            comp_sum = comparison_data['comparison_summary']

            # Title
            ws_comp_sum['A1'] = "Run Comparison Summary"
            ws_comp_sum['A1'].font = Font(bold=True, size=14)
            ws_comp_sum.merge_cells('A1:B1')

            row = 3

            # Summary metrics
            ws_comp_sum[f'A{row}'] = "COMPARISON METRICS"
            ws_comp_sum[f'A{row}'].font = section_font
            ws_comp_sum[f'A{row}'].fill = section_fill
            ws_comp_sum.merge_cells(f'A{row}:B{row}')
            row += 1

            # Data
            metrics = [
                ('Overall Trend', str(comp_sum.get('overall_trend', 'N/A')).title()),
                ('Total Error Delta', str(comp_sum.get('total_error_delta', 0))),
                ('Error Delta Percentage', f"{comp_sum.get('error_delta_percentage', 0):.1f}%"),
                ('Improved Steps', str(comp_sum.get('improved_steps', 0))),
                ('Degraded Steps', str(comp_sum.get('degraded_steps', 0))),
                ('Stable Steps', str(comp_sum.get('stable_steps', 0))),
                ('New Steps', str(comp_sum.get('new_steps', 0))),
                ('Removed Steps', str(comp_sum.get('removed_steps', 0)))
            ]

            for label, value in metrics:
                ws_comp_sum[f'A{row}'] = label
                ws_comp_sum[f'B{row}'] = value
                ws_comp_sum[f'A{row}'].font = Font(bold=True)
                row += 1

            auto_size_columns(ws_comp_sum)

        # ==================== SHEET 7: Step-by-Step Comparison ====================
        if 'step_comparisons' in comparison_data and comparison_data['step_comparisons']:
            ws_steps = wb.create_sheet("Step-by-Step Comparison")

            # Title
            ws_steps['A1'] = "Step-by-Step Comparison"
            ws_steps['A1'].font = Font(bold=True, size=14)
            ws_steps.merge_cells('A1:E1')

            row = 3

            # Headers
            ws_steps[f'A{row}'] = "Step Name"
            ws_steps[f'B{row}'] = "Run 1 Status"
            ws_steps[f'C{row}'] = "Run 2 Status"
            ws_steps[f'D{row}'] = "Trend"
            ws_steps[f'E{row}'] = "Change"
            style_header_row(ws_steps, row, 5)
            row += 1

            # Data
            for step in comparison_data['step_comparisons']:
                ws_steps[f'A{row}'] = step.get('step_name', 'N/A')
                ws_steps[f'B{row}'] = str(step.get('run1_status', 'N/A')) if step.get('exists_in_run1', True) else 'N/A'
                ws_steps[f'C{row}'] = str(step.get('run2_status', 'N/A')) if step.get('exists_in_run2', True) else 'N/A'
                ws_steps[f'D{row}'] = str(step.get('trend', 'N/A')).title()
                ws_steps[f'E{row}'] = str(step.get('change', 'N/A')).title()

                # Color code by trend/change
                change = step.get('change', '').lower()
                trend = step.get('trend', '').lower()

                if change == 'added':
                    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    ws_steps[f'E{row}'].fill = fill
                elif change == 'removed':
                    fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    ws_steps[f'E{row}'].fill = fill
                elif trend == 'improved':
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws_steps[f'D{row}'].fill = fill
                elif trend == 'degraded':
                    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws_steps[f'D{row}'].fill = fill

                row += 1

            auto_size_columns(ws_steps)

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"migration_readiness_report_{timestamp}.xlsx"

        # Return as downloadable file
        return StreamingResponse(
            iter([excel_buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export comparison Excel: {str(e)}")


@router.post("/export/pdf/comparison")
async def export_comparison_pdf(comparison_data: dict):
    """
    Export comparison data (with executive summary, trends, etc.) as comprehensive PDF report
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        import io

        pdf_buffer = io.BytesIO()

        # Create PDF document
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=1*inch,
            bottomMargin=0.75*inch
        )

        # Container for PDF elements
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=12,
            spaceBefore=12
        )
        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Heading3'],
            fontSize=13,
            textColor=colors.HexColor('#2E5C8A'),
            spaceAfter=8,
            spaceBefore=8
        )

        # Title
        elements.append(Paragraph("Migration Readiness Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # ==================== Executive Summary ====================
        if 'executive_summary' in comparison_data:
            elements.append(Paragraph("Executive Summary", heading_style))
            exec_sum = comparison_data['executive_summary']

            # Readiness Score Box
            score_data = [
                ['Readiness Score', str(exec_sum.get('readiness_score', 0))],
                ['Overall Status', exec_sum.get('overall_status', 'N/A')]
            ]
            score_table = Table(score_data, colWidths=[3*inch, 3*inch])
            score_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#D9E1F2')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 14),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.white)
            ]))
            elements.append(score_table)
            elements.append(Spacer(1, 15))

            # Validation Summary
            elements.append(Paragraph("Validation Summary", subheading_style))
            val_data = [
                ['Metric', 'Count'],
                ['Total Validations', str(exec_sum.get('total_validations', 0))],
                ['Passed', str(exec_sum.get('passed_validations', 0))],
                ['Failed', str(exec_sum.get('failed_validations', 0))],
                ['Warnings', str(exec_sum.get('warnings', 0))],
                ['Critical Issues', str(exec_sum.get('critical_issues', 0))]
            ]
            val_table = Table(val_data, colWidths=[3*inch, 3*inch])
            val_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(val_table)
            elements.append(Spacer(1, 15))

            # Severity Breakdown
            if 'severity_breakdown' in exec_sum:
                elements.append(Paragraph("Severity Breakdown", subheading_style))
                sev = exec_sum['severity_breakdown']
                sev_data = [
                    ['Severity', 'Count'],
                    ['BLOCKER', str(sev.get('BLOCKER', 0))],
                    ['HIGH', str(sev.get('HIGH', 0))],
                    ['MEDIUM', str(sev.get('MEDIUM', 0))],
                    ['LOW', str(sev.get('LOW', 0))]
                ]
                sev_table = Table(sev_data, colWidths=[3*inch, 3*inch])
                sev_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (0, 1), colors.red),
                    ('TEXTCOLOR', (0, 1), (0, 1), colors.white),
                    ('BACKGROUND', (0, 2), (0, 2), colors.orange),
                    ('BACKGROUND', (0, 3), (0, 3), colors.yellow),
                    ('BACKGROUND', (0, 4), (0, 4), colors.green),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey)
                ]))
                elements.append(sev_table)

            elements.append(PageBreak())

        # ==================== Comparison Summary ====================
        if 'comparison_summary' in comparison_data:
            elements.append(Paragraph("Comparison Summary", heading_style))
            comp_sum = comparison_data['comparison_summary']

            # Summary metrics table
            summary_data = [
                ['Metric', 'Value'],
                ['Overall Trend', str(comp_sum.get('overall_trend', 'N/A')).title()],
                ['Total Error Delta', str(comp_sum.get('total_error_delta', 0))],
                ['Error Delta %', f"{comp_sum.get('error_delta_percentage', 0):.1f}%"],
                ['Improved Steps', str(comp_sum.get('improved_steps', 0))],
                ['Degraded Steps', str(comp_sum.get('degraded_steps', 0))],
                ['Stable Steps', str(comp_sum.get('stable_steps', 0))],
                ['New Steps', str(comp_sum.get('new_steps', 0))],
                ['Removed Steps', str(comp_sum.get('removed_steps', 0))]
            ]
            summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 15))

        # ==================== Step-by-Step Comparison ====================
        if 'step_comparisons' in comparison_data and comparison_data['step_comparisons']:
            elements.append(Paragraph("Step-by-Step Comparison", heading_style))

            # Create table with step comparisons
            step_data = [['Step Name', 'Run 1', 'Run 2', 'Trend']]

            for step in comparison_data['step_comparisons'][:20]:  # Limit to 20 steps for PDF
                step_name = step.get('step_name', 'N/A')
                run1_status = step.get('run1_status', 'N/A')
                run2_status = step.get('run2_status', 'N/A')
                trend = step.get('trend', 'N/A')

                step_data.append([
                    Paragraph(step_name[:40], styles['Normal']) if len(step_name) > 40 else step_name,
                    str(run1_status),
                    str(run2_status),
                    str(trend).title()
                ])

            step_table = Table(step_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1*inch])
            step_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
            ]))
            elements.append(step_table)
            elements.append(PageBreak())

        # ==================== Trend Analysis ====================
        if 'trend_analysis' in comparison_data:
            elements.append(Paragraph("Trend Analysis", heading_style))
            trends = comparison_data['trend_analysis']

            # Velocity
            if 'velocity' in trends:
                elements.append(Paragraph("Velocity Metrics", subheading_style))
                vel = trends['velocity']
                vel_data = [
                    ['Metric', 'Value'],
                    ['Errors Resolved Per Day', str(vel.get('per_day', 0))],
                    ['Errors Resolved Per Week', str(vel.get('per_week', 0))]
                ]
                vel_table = Table(vel_data, colWidths=[3*inch, 3*inch])
                vel_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey)
                ]))
                elements.append(vel_table)
                elements.append(Spacer(1, 15))

            # Projections
            elements.append(Paragraph("Projections", subheading_style))
            proj_text = f"<b>Projected Zero Errors Date:</b> {trends.get('projected_zero_date', 'N/A')}<br/>"
            proj_text += f"<b>Regression Detected:</b> {trends.get('regression_detected', False)}"
            elements.append(Paragraph(proj_text, styles['Normal']))
            elements.append(Spacer(1, 15))

            # Error Trend
            if 'error_trend' in trends and trends['error_trend']:
                elements.append(Paragraph("Error Trend History", subheading_style))
                trend_data = [['Timestamp', 'Total Errors']]
                for point in trends['error_trend'][:20]:  # Limit to 20 most recent
                    trend_data.append([
                        point.get('timestamp', ''),
                        str(point.get('total_errors', 0))
                    ])

                trend_table = Table(trend_data, colWidths=[3.5*inch, 2.5*inch])
                trend_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('FONTSIZE', (0, 0), (-1, -1), 9)
                ]))
                elements.append(trend_table)

            elements.append(PageBreak())

        # ==================== Root Causes ====================
        if 'root_cause_groups' in comparison_data and comparison_data['root_cause_groups']:
            elements.append(Paragraph("Root Cause Analysis", heading_style))

            for idx, group in enumerate(comparison_data['root_cause_groups'][:15]):  # Limit to 15
                elements.append(Paragraph(f"{idx+1}. {group.get('title', 'Unknown')}", subheading_style))

                group_info = f"<b>Category:</b> {group.get('category', 'N/A')}<br/>"
                group_info += f"<b>Severity:</b> {group.get('severity', 'N/A')}<br/>"
                group_info += f"<b>Total Errors:</b> {group.get('total_errors', 0)}<br/>"
                group_info += f"<b>Affected Steps:</b> {group.get('total_affected', 0)}<br/>"
                group_info += f"<b>Description:</b> {group.get('description', 'N/A')}<br/>"
                group_info += f"<b>Recommended Action:</b> {group.get('recommended_action', 'N/A')}"

                elements.append(Paragraph(group_info, styles['Normal']))
                elements.append(Spacer(1, 10))

            elements.append(PageBreak())

        # ==================== Recommendations ====================
        if 'recommendations' in comparison_data and comparison_data['recommendations']:
            elements.append(Paragraph("Prioritized Recommendations", heading_style))

            for idx, rec in enumerate(comparison_data['recommendations'][:20]):  # Limit to 20
                priority = rec.get('priority', 'MEDIUM')
                elements.append(Paragraph(f"{idx+1}. [{priority}] {rec.get('title', 'Unknown')}", subheading_style))

                rec_info = f"<b>Category:</b> {rec.get('category', 'N/A')}<br/>"
                rec_info += f"<b>Effort:</b> {rec.get('effort', 'N/A')} | <b>Impact:</b> {rec.get('impact', 'N/A')}<br/>"
                rec_info += f"<b>Affected Count:</b> {rec.get('affected_count', 0)}<br/>"
                rec_info += f"<b>Description:</b> {rec.get('description', 'N/A')}<br/>"

                # Action items
                action_items = rec.get('action_items', [])
                if action_items:
                    rec_info += "<b>Action Items:</b><br/>"
                    for item in action_items[:5]:  # Limit to 5 action items
                        rec_info += f"&nbsp;&nbsp;• {item}<br/>"

                elements.append(Paragraph(rec_info, styles['Normal']))
                elements.append(Spacer(1, 10))

            if len(comparison_data['recommendations']) > 20:
                elements.append(Paragraph(f"<i>...and {len(comparison_data['recommendations']) - 20} more recommendations</i>", styles['Italic']))

        # Build PDF
        doc.build(elements)

        # Get PDF bytes
        pdf_buffer.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"migration_readiness_report_{timestamp}.pdf"

        # Return as downloadable file
        return StreamingResponse(
            iter([pdf_buffer.getvalue()]),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export comparison PDF: {str(e)}")